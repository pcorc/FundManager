from __future__ import annotations
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any, Dict, Iterable, Mapping, Optional

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from reporting.combined_reconciliation_report import build_combined_reconciliation_pdf
from reporting.report_utils import normalize_report_date


# ---------------------------------------------------------------------------
# Shared helpers
# ---------------------------------------------------------------------------
def _unwrap_if_date_keyed(
    payload: Mapping[str, Any],
    fund_level_keys: Iterable[str],
) -> Mapping[str, Any]:
    """If payload looks like {date: {fund: data}}, return the inner {fund: data}.

    Detection: the first nested value is a dict whose first value is itself a dict
    that contains at least one of ``fund_level_keys``. Otherwise the payload is
    returned unchanged.
    """
    if not payload:
        return payload
    first_value = next(iter(payload.values()))
    if not isinstance(first_value, dict) or not first_value:
        return payload
    sample = next(iter(first_value.values()))
    if isinstance(sample, dict) and any(k in sample for k in fund_level_keys):
        return first_value
    return payload


_NAV_FUND_LEVEL_KEYS = (
    "Beginning TNA",
    "Expected NAV",
    "Custodian NAV",
    "Expected TNA",
    "NAV Diff ($)",
    "summary",
    "detailed_calculations",
)


@dataclass
class GeneratedNAVReconciliationReport:
    """File locations for generated NAV reconciliation artefacts."""

    excel_path: Optional[str]


class NAVReconciliationExcelReport:
    """Render NAV reconciliation data into an Excel workbook with formulas and ticker details."""

    # Style + format constants — module-level constants would be just as fine,
    # but keeping them on the instance preserves the original API.
    CURRENCY_FORMAT = "#,##0.00"
    NAV_FORMAT = "#,##0.0000"
    INTEGER_FORMAT = "#,##0"
    PERCENT_FORMAT = "0.00%"

    def __init__(
        self,
        results: Mapping[str, Any],
        report_date: str,
        output_path: Path,
        workbook: Workbook | None = None,
    ) -> None:
        self.results = self._normalize_nav_payload(results)
        self.report_date = report_date
        self.output_path = output_path
        self.output_path.parent.mkdir(parents=True, exist_ok=True)
        self.workbook = workbook
        self._using_existing_workbook = workbook is not None

        self.title_font = Font(bold=True, size=14)
        self.header_font = Font(bold=True, size=12)
        self.subheader_font = Font(bold=True, size=10)
        self.body_font = Font(size=10)
        self.currency_format = self.CURRENCY_FORMAT
        self.nav_format = self.NAV_FORMAT
        self.integer_format = self.INTEGER_FORMAT
        self.percent_format = self.PERCENT_FORMAT

        self._sheet_refs: Dict[str, Dict[str, Any]] = {}

        self._export()

    # ------------------------------------------------------------------
    @staticmethod
    def _normalize_nav_payload(payload: Mapping[str, Any]) -> Dict[str, Any]:
        """Return fund-level data, unwrapping a date-keyed wrapper if present."""
        if not payload:
            return {}
        fund_level = _unwrap_if_date_keyed(payload, _NAV_FUND_LEVEL_KEYS)
        return {
            fund_name: fund_data
            for fund_name, fund_data in fund_level.items()
            if isinstance(fund_data, dict)
        }

    # ------------------------------------------------------------------
    def _export(self) -> None:
        workbook = self.workbook or Workbook()
        if not self._using_existing_workbook:
            # Drop the auto-created default sheet so we can position ours.
            workbook.remove(workbook.active)

        for fund_name, payload in sorted(self.results.items()):
            if payload:
                self._create_fund_sheet_with_formulas(workbook, fund_name, payload, self.report_date)

        summary_ws = self._create_summary_sheet(workbook)
        if summary_ws is not None and not self._using_existing_workbook:
            workbook._sheets.insert(0, workbook._sheets.pop(workbook._sheets.index(summary_ws)))

        workbook.save(self.output_path)

    # ------------------------------------------------------------------
    def _create_fund_sheet_with_formulas(self, wb, fund_name, nav_data, date_str):
        """Create detailed sheet for a fund with formulas matching NAVReconciliationReport."""
        ws = wb.create_sheet(fund_name[:31])
        self._sheet_refs[fund_name] = {"sheet_name": fund_name[:31], "components": {}}

        back = ws.cell(row=1, column=5, value="-> NAV Summary")
        back.hyperlink = "#'NAV Summary'!A1"
        back.font = Font(color="0000FF", underline="single")

        ws.merge_cells("A1:C1")
        ws["A1"] = f"{fund_name} - NAV Reconciliation - {date_str}"
        ws["A1"].font = self.title_font
        ws["A1"].alignment = Alignment(horizontal="center")

        row = 3
        ws.cell(row=row, column=1, value="NAV CALCULATION SUMMARY").font = self.header_font
        row += 2

        ws.cell(row=row, column=1, value="Beginning TNA (T-1)")
        ws.cell(row=row, column=3, value=nav_data.get("Beginning TNA", 0)).number_format = self.currency_format
        beg_tna_row = row
        row += 1

        ws.cell(row=row, column=1, value="Adjusted Beginning TNA")
        ws.cell(
            row=row, column=3,
            value=nav_data.get("Adjusted Beginning TNA", nav_data.get("Beginning TNA", 0)),
        ).number_format = self.currency_format
        adj_beg_tna_row = row
        row += 2

        ws.cell(row=row, column=1, value="Gain/Loss Components:").font = self.subheader_font
        row += 1

        component_row_refs: Dict[str, str] = {}

        # Detail sections start below the summary block.
        detail_start_row = row + 30

        equity_info = self._add_asset_details_with_formulas(
            ws, detail_start_row, nav_data,
            title="EQUITY GAIN/LOSS DETAIL",
            details_key="equity_details",
            empty_message="No equity holdings with gain/loss",
        )
        option_info = self._add_asset_details_with_formulas(
            ws, equity_info["end_row"] + 3, nav_data,
            title="OPTION GAIN/LOSS DETAIL",
            details_key="option_details",
            empty_message="No option holdings with gain/loss",
            contract_multiplier=100,
        )
        flex_info = self._add_asset_details_with_formulas(
            ws, option_info["end_row"] + 3, nav_data,
            title="FLEX OPTION GAIN/LOSS DETAIL",
            details_key="flex_details",
            empty_message="No flex option holdings with gain/loss",
            contract_multiplier=100,
        )

        # Equity G/L
        ws.cell(row=row, column=1, value="  Equity G/L")
        ws.cell(row=row, column=3, value=(
            f"={equity_info['gl_total_cell']}" if equity_info["gl_total_cell"]
            else nav_data.get("Equity G/L", 0)
        )).number_format = self.currency_format
        component_row_refs["equity"] = f"C{row}"
        self._sheet_refs[fund_name]["components"]["equity"] = f"C{row}"
        row += 1

        ws.cell(row=row, column=1, value="  Equity G/L (Adjusted)")
        ws.cell(row=row, column=3, value=(
            f"={equity_info['gl_adj_total_cell']}" if equity_info["gl_adj_total_cell"]
            else nav_data.get("Equity G/L Adj", nav_data.get("Equity G/L", 0))
        )).number_format = self.currency_format
        row += 1

        # Option G/L
        option_gl = nav_data.get("Option G/L", 0)
        if abs(option_gl) > 0.01 or option_info["has_data"]:
            ws.cell(row=row, column=1, value="  Option G/L")
            ws.cell(row=row, column=3, value=(
                f"={option_info['gl_total_cell']}" if option_info["gl_total_cell"]
                else option_gl
            )).number_format = self.currency_format
            component_row_refs["option"] = f"C{row}"
            self._sheet_refs[fund_name]["components"]["options"] = f"C{row}"
            row += 1

            ws.cell(row=row, column=1, value="  Option G/L (Adjusted)")
            ws.cell(row=row, column=3, value=(
                f"={option_info['gl_adj_total_cell']}" if option_info["gl_adj_total_cell"]
                else nav_data.get("Option G/L Adj", option_gl)
            )).number_format = self.currency_format
            row += 1

        # Flex Option G/L
        flex_gl = nav_data.get("Flex Option G/L", 0)
        if abs(flex_gl) > 0.01 or flex_info["has_data"]:
            ws.cell(row=row, column=1, value="  Flex Option G/L")
            ws.cell(row=row, column=3, value=(
                f"={flex_info['gl_total_cell']}" if flex_info["gl_total_cell"]
                else flex_gl
            )).number_format = self.currency_format
            component_row_refs["flex"] = f"C{row}"
            self._sheet_refs[fund_name]["components"]["flex_options"] = f"C{row}"
            row += 1

            ws.cell(row=row, column=1, value="  Flex Option G/L (Adjusted)")
            ws.cell(row=row, column=3, value=(
                f"={flex_info['gl_adj_total_cell']}" if flex_info["gl_adj_total_cell"]
                else nav_data.get("Flex Option G/L Adj", flex_gl)
            )).number_format = self.currency_format
            row += 1

        # Treasury G/L
        treasury_gl = nav_data.get("Treasury G/L", 0)
        if abs(treasury_gl) > 0.01:
            ws.cell(row=row, column=1, value="  Treasury G/L")
            ws.cell(row=row, column=3, value=treasury_gl).number_format = self.currency_format
            component_row_refs["treasury"] = f"C{row}"
            self._sheet_refs[fund_name]["components"]["treasury"] = f"C{row}"
            row += 1

        # Assignment G/L
        assignment_gl = nav_data.get("Assignment G/L", 0)
        if abs(assignment_gl) > 0.01:
            ws.cell(row=row, column=1, value="  Assignment G/L")
            ws.cell(row=row, column=3, value=assignment_gl).number_format = self.currency_format
            row += 1

        # Other
        other = nav_data.get("Other", 0)
        if abs(other) > 0.01:
            ws.cell(row=row, column=1, value="  Other")
            ws.cell(row=row, column=3, value=other).number_format = self.currency_format
            row += 1

        row += 1

        # Expenses
        ws.cell(row=row, column=1, value="Expenses")
        expenses_value = -abs(nav_data.get("Accruals", 0))
        ws.cell(row=row, column=3, value=expenses_value).number_format = self.currency_format
        expenses_row = row
        self._sheet_refs[fund_name]["expenses"] = f"C{row}"
        row += 1

        # Dividends
        dividends = nav_data.get("Dividends", 0)
        dividends_row = None
        if abs(dividends) > 0.01:
            ws.cell(row=row, column=1, value="Dividends")
            ws.cell(row=row, column=3, value=dividends).number_format = self.currency_format
            dividends_row = row
            self._sheet_refs[fund_name]["dividends"] = f"C{row}"
            row += 1

        # Distributions
        distributions = nav_data.get("Distributions", 0)
        distributions_row = None
        if abs(distributions) > 0.01:
            ws.cell(row=row, column=1, value="Distributions")
            ws.cell(row=row, column=3, value=-abs(distributions)).number_format = self.currency_format
            distributions_row = row
            self._sheet_refs[fund_name]["distributions"] = f"C{row}"
            row += 1

        row += 1

        # Expected TNA formula
        ws.cell(row=row, column=1, value="Expected TNA").font = self.subheader_font
        formula_parts = [f"C{adj_beg_tna_row}"]
        for adj_cell in (
            equity_info.get("gl_adj_total_cell"),
            option_info.get("gl_adj_total_cell"),
            flex_info.get("gl_adj_total_cell"),
        ):
            if adj_cell:
                formula_parts.append(adj_cell)
        if "treasury" in component_row_refs:
            formula_parts.append(component_row_refs["treasury"])
        if abs(assignment_gl) > 0.01:
            formula_parts.append(str(nav_data.get("Assignment G/L", 0)))
        if abs(other) > 0.01:
            formula_parts.append(str(other))
        formula_parts.append(f"C{expenses_row}")
        if dividends_row:
            formula_parts.append(f"C{dividends_row}")
        if distributions_row:
            formula_parts.append(f"C{distributions_row}")
        ws.cell(row=row, column=3, value="=" + "+".join(formula_parts)).number_format = self.currency_format
        expected_tna_row = row
        self._sheet_refs[fund_name]["expected_tna"] = f"C{row}"
        row += 1

        # Custodian TNA
        ws.cell(row=row, column=1, value="Custodian TNA").font = self.subheader_font
        ws.cell(row=row, column=3, value=nav_data.get("Custodian TNA", 0)).number_format = self.currency_format
        cust_tna_row = row
        self._sheet_refs[fund_name]["cust_tna"] = f"C{row}"
        row += 1

        # TNA Difference
        ws.cell(row=row, column=1, value="TNA Difference")
        ws.cell(row=row, column=3, value=f"=C{cust_tna_row}-C{expected_tna_row}").number_format = self.currency_format
        row += 2

        # Shares Outstanding
        ws.cell(row=row, column=1, value="Shares Outstanding")
        ws.cell(row=row, column=3, value=nav_data.get("Shares Outstanding", 0)).number_format = "#,##0"
        shares_row = row
        row += 2

        # Expected NAV
        ws.cell(row=row, column=1, value="Expected NAV").font = self.subheader_font
        ws.cell(
            row=row, column=3,
            value=f"=IF(C{shares_row}<>0,C{expected_tna_row}/C{shares_row},0)",
        ).number_format = self.nav_format
        expected_nav_row = row
        self._sheet_refs[fund_name]["expected_nav"] = f"C{row}"
        row += 1

        # Custodian NAV
        ws.cell(row=row, column=1, value="Custodian NAV").font = self.subheader_font
        ws.cell(row=row, column=3, value=nav_data.get("Custodian NAV", 0)).number_format = self.nav_format
        cust_nav_row = row
        self._sheet_refs[fund_name]["cust_nav"] = f"C{row}"
        row += 1

        # NAV Difference
        ws.cell(row=row, column=1, value="NAV Difference")
        ws.cell(row=row, column=3, value=f"=C{cust_nav_row}-C{expected_nav_row}").number_format = self.nav_format
        nav_diff_row = row
        self._sheet_refs[fund_name]["variance"] = f"C{row}"
        row += 2

        # NAV Good indicators
        ws.cell(row=row, column=1, value="NAV Good (2 decimal)")
        ws.cell(row=row, column=3, value=f'=IF(ABS(ROUND(C{nav_diff_row},2))<=0.01,"PASS","FAIL")')
        row += 1

        ws.cell(row=row, column=1, value="NAV Good (4 decimal)")
        ws.cell(row=row, column=3, value=f'=IF(ABS(ROUND(C{nav_diff_row},4))<=0.0001,"PASS","FAIL")')

        self._sheet_refs[fund_name]["prior_nav"] = f"C{beg_tna_row}"

        for col in range(1, 12):
            ws.column_dimensions[get_column_letter(col)].width = 14
        ws.freeze_panes = "A4"

    # ------------------------------------------------------------------
    def _add_asset_details_with_formulas(
        self,
        ws,
        start_row: int,
        nav_data: Mapping[str, Any],
        *,
        title: str,
        details_key: str,
        empty_message: str,
        contract_multiplier: int = 1,
    ) -> Dict[str, Any]:
        """Render a per-ticker gain/loss section. Used for equity, options, and flex options.

        ``contract_multiplier`` is the factor applied to the G/L formula:
        1 for cash equities, 100 for options/flex options (per-contract underlying).
        """
        ws.cell(row=start_row, column=1, value=title).font = self.header_font

        details_df = nav_data.get("detailed_calculations", {}).get(details_key, pd.DataFrame())
        if details_df is None or details_df.empty:
            ws.cell(row=start_row + 2, column=1, value=empty_message)
            return {
                "has_data": False,
                "gl_total_cell": None,
                "gl_adj_total_cell": None,
                "end_row": start_row + 2,
            }

        headers = [
            "Ticker", "Qty T-1", "Qty T",
            "Price T-1\n(Vest)", "Price T\n(Raw)",
            "Price T-1\n(Cust)", "Price T\n(Adj)",
            "G/L", "G/L Adj",
        ]
        header_row = start_row + 2
        for i, header in enumerate(headers):
            cell = ws.cell(row=header_row, column=i + 1, value=header)
            cell.font = self.subheader_font
            cell.alignment = Alignment(horizontal="center", wrap_text=True)

        first_data_row = header_row + 1
        data_row = first_data_row
        multiplier_suffix = f"*{contract_multiplier}" if contract_multiplier != 1 else ""

        for _, row_data in details_df.iterrows():
            ws.cell(row=data_row, column=1, value=row_data.get("ticker", ""))
            ws.cell(row=data_row, column=2, value=row_data.get("quantity_t1", 0)).number_format = "#,##0"
            ws.cell(row=data_row, column=3, value=row_data.get("quantity_t", 0)).number_format = "#,##0"

            price_t1_raw = row_data.get("price_t1_raw", 0)
            price_t_raw = row_data.get("price_t_raw", 0)
            ws.cell(row=data_row, column=4, value=price_t1_raw).number_format = self.currency_format
            ws.cell(row=data_row, column=5, value=price_t_raw).number_format = self.currency_format

            price_t1_adj = row_data.get("price_t1_adj", price_t1_raw)
            price_t_adj = row_data.get("price_t_adj", price_t_raw)

            adj_t1_cell = ws.cell(row=data_row, column=6, value=price_t1_adj)
            adj_t1_cell.number_format = self.currency_format
            if abs(price_t1_adj - price_t1_raw) > 0.001:
                adj_t1_cell.fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")

            adj_t_cell = ws.cell(row=data_row, column=7, value=price_t_adj)
            adj_t_cell.number_format = self.currency_format
            if abs(price_t_adj - price_t_raw) > 0.001:
                adj_t_cell.fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")

            ws.cell(
                row=data_row, column=8,
                value=f"=(E{data_row}-F{data_row})*C{data_row}{multiplier_suffix}",
            ).number_format = self.currency_format
            ws.cell(
                row=data_row, column=9,
                value=f"=(G{data_row}-F{data_row})*C{data_row}{multiplier_suffix}",
            ).number_format = self.currency_format

            data_row += 1

        if data_row == first_data_row:
            return {
                "has_data": False,
                "gl_total_cell": None,
                "gl_adj_total_cell": None,
                "end_row": data_row,
            }

        last_data_row = data_row - 1
        total_row = data_row + 1

        ws.cell(row=total_row, column=1, value="TOTAL").font = self.subheader_font
        gl_range = f"H{first_data_row}:H{last_data_row}"
        gl_adj_range = f"I{first_data_row}:I{last_data_row}"
        ws.cell(row=total_row, column=8, value=f"=SUM({gl_range})").number_format = self.currency_format
        ws.cell(row=total_row, column=9, value=f"=SUM({gl_adj_range})").number_format = self.currency_format

        return {
            "has_data": True,
            "gl_total_cell": f"H{total_row}",
            "gl_adj_total_cell": f"I{total_row}",
            "end_row": total_row,
        }

    def _create_summary_sheet(self, workbook: Workbook):
        """Create summary sheet with all funds. Column A links to each fund's tab."""
        worksheet = workbook.create_sheet("NAV Summary")

        worksheet.merge_cells("A1:K1")
        title_cell = worksheet.cell(
            row=1, column=1, value=f"NAV Reconciliation Summary - {self.report_date}"
        )
        title_cell.font = self.title_font
        title_cell.alignment = Alignment(horizontal="center")

        headers = [
            "Fund", "Date", "Expected TNA", "Custodian TNA", "TNA Diff ($)",
            "Expected NAV", "Custodian NAV", "NAV Diff ($)", "Shares Outstanding",
            "NAV Good (2-dec)", "NAV Good (4-dec)",
        ]
        header_row = 3
        for index, header in enumerate(headers, start=1):
            cell = worksheet.cell(row=header_row, column=index, value=header)
            cell.font = self.header_font
            cell.alignment = Alignment(horizontal="center")

        good_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
        bad_fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
        link_font = Font(color="0000FF", underline="single")

        current_row = header_row + 1
        for fund_name in sorted(self.results.keys()):
            nav_data = self.results[fund_name]

            # Fund cell with an internal hyperlink to that fund's tab.
            fund_cell = worksheet.cell(row=current_row, column=1, value=fund_name)
            sheet_name = self._sheet_refs.get(fund_name, {}).get("sheet_name", fund_name[:31])
            # Quote the sheet name so names with spaces/symbols still resolve.
            fund_cell.hyperlink = f"#'{sheet_name}'!A1"
            fund_cell.font = link_font

            worksheet.cell(row=current_row, column=2, value=self.report_date)
            worksheet.cell(row=current_row, column=3, value=nav_data.get("Expected TNA", 0)).number_format = self.currency_format
            worksheet.cell(row=current_row, column=4, value=nav_data.get("Custodian TNA", 0)).number_format = self.currency_format
            worksheet.cell(row=current_row, column=5, value=nav_data.get("TNA Diff ($)", 0)).number_format = self.currency_format
            worksheet.cell(row=current_row, column=6, value=nav_data.get("Expected NAV", 0)).number_format = self.nav_format
            worksheet.cell(row=current_row, column=7, value=nav_data.get("Custodian NAV", 0)).number_format = self.nav_format
            worksheet.cell(row=current_row, column=8, value=nav_data.get("NAV Diff ($)", 0)).number_format = self.nav_format
            worksheet.cell(row=current_row, column=9, value=nav_data.get("Shares Outstanding", 0)).number_format = "#,##0"

            nav_good_2 = bool(nav_data.get("NAV Good (2 Digit)", False))
            nav_good_4 = bool(nav_data.get("NAV Good (4 Digit)", False))

            cell_2dec = worksheet.cell(row=current_row, column=10, value="PASS" if nav_good_2 else "FAIL")
            cell_4dec = worksheet.cell(row=current_row, column=11, value="PASS" if nav_good_4 else "FAIL")
            cell_2dec.fill = good_fill if nav_good_2 else bad_fill
            cell_4dec.fill = good_fill if nav_good_4 else bad_fill

            current_row += 1

        for column_index in range(1, 12):
            worksheet.column_dimensions[get_column_letter(column_index)].width = 15

        return worksheet

# ------------------------------------------------------------------------
def convert_nav_results_to_dicts(nav_data: Mapping[str, Any]) -> Dict[str, Any]:
    """Recursively convert NAVReconciliationResults dataclasses to plain dicts."""
    converted: Dict[str, Any] = {}
    for key, value in nav_data.items():
        if hasattr(value, "to_legacy_dict"):
            converted[key] = value.to_legacy_dict()
        elif isinstance(value, dict):
            converted[key] = convert_nav_results_to_dicts(value)
        else:
            converted[key] = value
    return converted


def generate_nav_reconciliation_reports(
    reconciliation_results,
    date_str,
    excel_path,
    *,
    workbook: Workbook | None = None,
):
    """Generate NAV reconciliation reports in Excel format."""
    resolved_date = normalize_report_date(date_str)
    converted_results = convert_nav_results_to_dicts(reconciliation_results)

    if isinstance(converted_results, dict):
        if date_str in converted_results:
            normalized = converted_results[date_str]
        elif resolved_date in converted_results:
            normalized = converted_results[resolved_date]
        else:
            normalized = _unwrap_if_date_keyed(converted_results, _NAV_FUND_LEVEL_KEYS)
    else:
        normalized = {}

    output_path = Path(excel_path)
    if output_path.suffix.lower() != ".xlsx":
        output_path = output_path / f"nav_reconciliation_{resolved_date}.xlsx"

    if workbook is None and output_path.exists():
        workbook = load_workbook(output_path)

    NAVReconciliationExcelReport(
        results=normalized,
        report_date=resolved_date,
        output_path=output_path,
        workbook=workbook,
    )

    return str(output_path)


_HOLDINGS_FUND_LEVEL_KEYS = ("summary", "details", "custodian_equity")


def generate_reconciliation_summary_pdf(
    reconciliation_results: Mapping[str, Any],
    nav_results: Mapping[str, Any],
    report_date: date | datetime | str,
    output_dir: str,
    *,
    file_name: str | None = None,
) -> Optional[str]:
    """Create the combined holdings + NAV reconciliation PDF."""

    date_str = normalize_report_date(report_date)
    raw_recon = _unwrap_if_date_keyed(reconciliation_results or {}, _HOLDINGS_FUND_LEVEL_KEYS)
    raw_nav = _unwrap_if_date_keyed(
        convert_nav_results_to_dicts(nav_results or {}),
        _NAV_FUND_LEVEL_KEYS,
    )

    if not raw_recon and not raw_nav:
        return None

    output_path = Path(output_dir)
    output_path.mkdir(parents=True, exist_ok=True)

    resolved_name = file_name or f"reconciliation_summary_{date_str}.pdf"
    pdf_path = output_path / resolved_name

    holdings_details: Dict[str, Dict[str, Any]] = {date_str: {}}
    holdings_summary: Dict[str, list[Dict[str, Any]]] = {date_str: []}
    for fund, payload in raw_recon.items():
        if not payload:
            continue
        holdings_details[date_str][fund] = payload.get("details", {})
        holdings_summary[date_str].append({"fund": fund, "summary": payload.get("summary", {})})

    nav_details: Dict[str, Dict[str, Any]] = {date_str: {}}
    nav_summary: Dict[str, list[Dict[str, Any]]] = {date_str: []}
    for fund, payload in raw_nav.items():
        if not payload:
            continue
        nav_details[date_str][fund] = payload
        nav_summary[date_str].append({"fund": fund, "summary": payload})

    build_combined_reconciliation_pdf(
        nav_details,
        holdings_details,
        nav_summary,
        holdings_summary,
        date_str,
        pdf_path,
    )

    return str(pdf_path)
