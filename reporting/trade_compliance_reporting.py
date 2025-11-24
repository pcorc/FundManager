"""Generate Excel and PDF outputs for trading compliance comparisons."""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, Iterable, Mapping, Optional, Sequence, Tuple

import logging

from fpdf import FPDF
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from reporting.report_utils import format_number, normalize_report_date

logger = logging.getLogger(__name__)

ASSET_KEY_ALIASES: Dict[str, Tuple[str, ...]] = {
    "equity": ("equity", "equities"),
    "options": ("options", "option"),
    "treasury": ("treasury", "treasuries", "bonds", "bond"),
}

ASSET_LABELS: Dict[str, str] = {
    "equity": "Equity",
    "options": "Options",
    "treasury": "Treasury",
}

ASSET_ORDER: Tuple[str, ...] = ("equity", "options", "treasury")


def _asset_key_candidates(asset_key: str) -> Tuple[str, ...]:
    aliases = ASSET_KEY_ALIASES.get(asset_key, (asset_key,))
    if asset_key in aliases:
        ordered: Iterable[str] = aliases
    else:
        ordered = (asset_key,) + aliases
    seen: list[str] = []
    for candidate in ordered:
        if candidate not in seen:
            seen.append(candidate)
    return tuple(seen) if seen else (asset_key,)


def _extract_asset_mapping(source: Mapping[str, Any], asset_key: str) -> Dict[str, Any]:
    if not isinstance(source, Mapping):
        return {}
    for candidate in _asset_key_candidates(asset_key):
        value = source.get(candidate)
        if isinstance(value, Mapping):
            return dict(value)
    return {}


def _build_asset_sections(trade_info: Mapping[str, Any]) -> Sequence[Dict[str, Any]]:
    asset_summary = trade_info.get("asset_trade_summary", {}) or {}
    net_trades = trade_info.get("net_trades", {}) or {}
    activity_info = trade_info.get("trade_activity", {}) or {}

    sections: list[Dict[str, Any]] = []
    for asset_key in ASSET_ORDER:
        summary = _extract_asset_mapping(asset_summary, asset_key)
        activity_details = _extract_asset_mapping(activity_info, asset_key)

        net_payload: Dict[str, Any] = {}
        if isinstance(activity_details, Mapping):
            net_payload.update(dict(activity_details.get("net", {}) or {}))
        for candidate in _asset_key_candidates(asset_key):
            candidate_net = net_trades.get(candidate)
            if isinstance(candidate_net, Mapping):
                if not net_payload:
                    net_payload = dict(candidate_net)
                else:
                    net_payload.update(candidate_net)
                break

        sections.append(
            {
                "asset_key": asset_key,
                "label": ASSET_LABELS.get(asset_key, asset_key.title()),
                "summary": summary,
                "activity": activity_details if isinstance(activity_details, Mapping) else {},
                "net": net_payload,
            }
        )

    return sections


@dataclass
class GeneratedTradingComplianceReport:
    """Excel/PDF artefacts for trading compliance."""

    excel_path: Optional[str]
    pdf_path: Optional[str]


def _format_percent(value: float) -> str:
    try:
        return f"{float(value) * 100:.2f}%"
    except (TypeError, ValueError):
        return "0.00%"

def generate_trading_excel_report(comparison_data: Mapping[str, Any], output_path: str) -> None:
    """Create a multi-tab Excel workbook summarising trading compliance changes."""

    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    _create_summary_sheet(workbook, comparison_data)
    _create_trade_activity_sheet(workbook, comparison_data)
    _create_compliance_details_sheet(workbook, comparison_data)

    workbook.save(path)
    logger.info("Trading comparison report saved to: %s", path)


def _create_summary_sheet(workbook: Workbook, data: Mapping[str, Any]) -> None:
    """Create a compliance overview sheet with key status changes."""

    summary = data.get("summary", {}) or {}

    sheet = workbook.create_sheet("Compliance Overview", 0)
    sheet["A1"] = "Trading Compliance Analysis - Compliance Overview"
    sheet["A1"].font = Font(size=14, bold=True)
    sheet["A2"] = f"Date: {data.get('date', '')}"

    row = 4
    row = _append_compliance_changes_table(sheet, data, summary, start_row=row)
    _append_fund_summary_metrics_table(sheet, data, start_row=row + 2)

def _update_width(width_tracker: Dict[int, int], col_idx: int, value: object) -> None:
    """Update the width tracker for a column with the length of a value."""
    text = "" if value is None else str(value)
    current = width_tracker.get(col_idx, 0)
    width_tracker[col_idx] = max(current, len(text))

def _append_fund_summary_metrics_table(
    sheet, data: Mapping[str, Any], *, start_row: int
) -> int:
    funds = data.get("funds", {}) or {}

    metric_labels = {
        "cash_value": "Cash Value",
        "equity_market_value": "Equity Market Value",
        "option_market_value": "Option Market Value",
        "option_delta_adjusted_notional": "Option Delta Adjusted Notional",
        "treasury": "Treasury Market Value",
        "total_assets": "Total Assets",
        "total_net_assets": "Total Net Assets",
    }

    populated = False
    row = start_row
    width_tracker: Dict[int, int] = {}  # Initialize width tracker

    for fund_name, fund_data in sorted(funds.items()):
        summary_metrics = fund_data.get("summary_metrics", {}) or {}
        ex_ante = dict(summary_metrics.get("ex_ante", {}) or {})
        ex_post = dict(summary_metrics.get("ex_post", {}) or {})

        if not ex_ante and not ex_post:
            continue

        # Update width for the section header
        _update_width(width_tracker, 1, "Fund Summary Metrics")
        populated = True

        sheet[f"A{row}"] = f"Fund Summary Metrics - {fund_name}"
        sheet[f"A{row}"].font = Font(size=11, bold=True)
        _update_width(width_tracker, 1, f"Fund Summary Metrics - {fund_name}")
        row += 1

        headers = ["Metric", "Ex-Ante", "Ex-Post", "Delta"]
        for col, header in enumerate(headers, start=1):
            cell = sheet.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(
                start_color="DDDDDD", end_color="DDDDDD", fill_type="solid"
            )
            cell.alignment = Alignment(horizontal="center")
            _update_width(width_tracker, col, header)
        row += 1

        wrote_row = False
        for metric_key, label in metric_labels.items():
            if metric_key not in ex_ante and metric_key not in ex_post:
                continue

            ante_value = float(ex_ante.get(metric_key, 0.0) or 0.0)
            post_value = float(ex_post.get(metric_key, 0.0) or 0.0)
            delta = post_value - ante_value

            sheet.cell(row=row, column=1, value=label)
            ante_cell = sheet.cell(row=row, column=2, value=ante_value)
            ante_cell.number_format = "#,##0.00"
            ante_cell.alignment = Alignment(horizontal="right")
            post_cell = sheet.cell(row=row, column=3, value=post_value)
            post_cell.number_format = "#,##0.00"
            post_cell.alignment = Alignment(horizontal="right")
            delta_cell = sheet.cell(row=row, column=4, value=delta)
            delta_cell.number_format = "#,##0.00"
            delta_cell.alignment = Alignment(horizontal="right")

            _update_width(width_tracker, 1, label)
            _update_width(width_tracker, 2, f"{ante_value:,.2f}")
            _update_width(width_tracker, 3, f"{post_value:,.2f}")
            _update_width(width_tracker, 4, f"{delta:,.2f}")

            row += 1
            wrote_row = True

        if wrote_row:
            row += 2

    # Apply the calculated column widths
    if populated:
        for col_idx, width in width_tracker.items():
            column = get_column_letter(col_idx)
            adjusted_width = min(max(width + 2, 14), 60)
            sheet.column_dimensions[column].width = adjusted_width

    return row

def _append_compliance_changes_table(
    sheet, data: Mapping[str, Any], summary: Mapping[str, Any], *, start_row: int
) -> int:
    funds = data.get("funds", {}) or {}

    row = start_row
    sheet[f"A{row}"] = "Compliance Status Changes"
    sheet[f"A{row}"].font = Font(size=12, bold=True)
    row += 1

    sheet[f"A{row}"] = "Total Funds Analyzed"
    sheet[f"A{row}"].font = Font(bold=True)
    sheet[f"B{row}"] = summary.get("total_funds_analyzed", 0)
    sheet[f"B{row}"].font = Font(bold=True)
    sheet[f"B{row}"].alignment = Alignment(horizontal="center")
    row += 1

    sheet[f"A{row}"] = "Total Funds Traded"
    sheet[f"A{row}"].font = Font(bold=True)
    sheet[f"B{row}"] = summary.get("total_funds_traded", 0)
    sheet[f"B{row}"].font = Font(bold=True)
    sheet[f"B{row}"].alignment = Alignment(horizontal="center")
    row += 2

    headers = [
        "Fund Name",
        "Status Change",
        "Violations Before",
        "Violations After",
        "Net Change",
    ]

    for col, header in enumerate(headers, start=1):
        cell = sheet.cell(row=row, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    row += 1

    total_before = 0
    total_after = 0
    total_net = 0

    for fund_name, fund_data in sorted(funds.items()):
        violations_before = int(fund_data.get("violations_before", 0) or 0)
        violations_after = int(fund_data.get("violations_after", 0) or 0)
        net_change = violations_after - violations_before

        total_before += violations_before
        total_after += violations_after
        total_net += net_change

        sheet.cell(row=row, column=1, value=fund_name)
        status_cell = sheet.cell(
            row=row, column=2, value=fund_data.get("status_change", "UNCHANGED")
        )
        status_cell.alignment = Alignment(horizontal="center")
        before_cell = sheet.cell(row=row, column=3, value=violations_before)
        before_cell.alignment = Alignment(horizontal="center")
        after_cell = sheet.cell(row=row, column=4, value=violations_after)
        after_cell.alignment = Alignment(horizontal="center")
        net_cell = sheet.cell(row=row, column=5, value=net_change)
        net_cell.alignment = Alignment(horizontal="center")

        status = (fund_data.get("status_change") or "UNCHANGED").upper()
        fill_color: Optional[str]
        if status == "INTO_COMPLIANCE":
            fill_color = "CCFFCC"
        elif status == "OUT_OF_COMPLIANCE":
            fill_color = "FFE0B3"
        elif status == "UNCHANGED" and violations_after > 0:
            fill_color = "FF9999"
        else:
            fill_color = None

        if fill_color:
            fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
            if fill_color == "FF9999":
                sheet.cell(row=row, column=1).fill = fill
            else:
                for col_idx in range(1, len(headers) + 1):
                    sheet.cell(row=row, column=col_idx).fill = fill

        row += 1

    totals_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    totals_font = Font(bold=True)
    labels = ["Totals", "", total_before, total_after, total_net]
    for col_idx, value in enumerate(labels, start=1):
        cell = sheet.cell(row=row, column=col_idx, value=value)
        cell.font = totals_font
        cell.fill = totals_fill
        if col_idx >= 2:
            cell.alignment = Alignment(horizontal="center")
    row += 1

    for col in range(1, len(headers) + 1):
        column = get_column_letter(col)
        current_width = sheet.column_dimensions[column].width
        min_width = 28 if col == 1 else 22
        if current_width is None or current_width < min_width:
            sheet.column_dimensions[column].width = min_width

    return row

def _create_trade_activity_sheet(workbook: Workbook, data: Mapping[str, Any]) -> None:
    """Tab describing detailed trading activity by fund and asset class."""

    sheet = workbook.create_sheet("Trade Activity")
    sheet["A1"] = "Trade Activity Overview"
    sheet["A1"].font = Font(size=12, bold=True)
    sheet.freeze_panes = "A3"

    summary_headers = [
        "Fund",
        "Asset Class",
        "Total Net Assets",
        "Total Assets",
        "Trade Value",
        "% of TNA",
        "% of Total Assets",
        "Ex-Ante Market Value",
        "Ex-Post Market Value",
        "Market Value Delta",
        "Trade vs Ex-Ante %",
        "Ex-Post vs Ex-Ante %",
        "Net Trade Value",
        "Net Buy Qty",
        "Net Sell Qty",
        "Net Buy Value",
        "Net Sell Value",
    ]

    header_row = 3
    for col, header in enumerate(summary_headers, start=1):
        cell = sheet.cell(row=header_row, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    row = header_row + 1
    summary_rows_written = False

    for fund_name, fund_data in sorted(data.get("funds", {}).items()):
        trade_info = fund_data.get("trade_info", {}) or {}
        sections = _build_asset_sections(trade_info)

        total_net_assets = float(trade_info.get("total_net_assets", 0.0) or 0.0)
        total_assets = float(trade_info.get("total_assets", 0.0) or 0.0)

        for section in sections:
            summary = section.get("summary", {}) or {}
            activity_details = section.get("activity", {}) or {}
            net_info = section.get("net", {}) or {}

            trade_value = float(summary.get("trade_value", 0.0) or 0.0)
            market_delta = float(summary.get("market_value_delta", 0.0) or 0.0)
            pct_tna = float(summary.get("pct_of_tna", 0.0) or 0.0)
            pct_assets = float(summary.get("pct_of_total_assets", 0.0) or 0.0)
            ex_ante = float(summary.get("ex_ante_market_value", 0.0) or 0.0)
            ex_post = float(summary.get("ex_post_market_value", 0.0) or 0.0)
            trade_vs_ante = float(summary.get("trade_vs_ex_ante_pct", 0.0) or 0.0)
            post_vs_ante = float(summary.get("ex_post_vs_ex_ante_pct", 0.0) or 0.0)
            buy_qty = float(net_info.get("buy_quantity", net_info.get("buys", 0.0)) or 0.0)
            sell_qty = float(net_info.get("sell_quantity", net_info.get("sells", 0.0)) or 0.0)
            buy_val = float(net_info.get("buy_value", 0.0) or 0.0)
            sell_val = float(net_info.get("sell_value", 0.0) or 0.0)

            if trade_value == 0.0:
                trade_value = abs(buy_val) + abs(sell_val)

            raw_net_value = net_info.get("net_value")
            if raw_net_value in (None, ""):
                net_value = buy_val - sell_val
            else:
                try:
                    net_value = float(raw_net_value)
                except (TypeError, ValueError):
                    net_value = buy_val - sell_val

            buys_list = activity_details.get("buys") if isinstance(activity_details, Mapping) else []
            sells_list = activity_details.get("sells") if isinstance(activity_details, Mapping) else []
            has_trade_lists = bool(buys_list or sells_list)

            if not has_trade_lists and not any(
                    abs(val) > 0.0
                    for val in (
                            trade_value,
                            market_delta,
                            net_value,
                            buy_qty,
                            sell_qty,
                            buy_val,
                            sell_val,
                    )
            ):
                continue

            summary_rows_written = True

            sheet.cell(row=row, column=1, value=fund_name)
            sheet.cell(row=row, column=2, value=section.get("label", section.get("asset_key", "")).title())

            tna_cell = sheet.cell(row=row, column=3, value=total_net_assets)
            tna_cell.number_format = "#,##0.00"

            total_assets_cell = sheet.cell(row=row, column=4, value=total_assets)
            total_assets_cell.number_format = "#,##0.00"

            trade_val_cell = sheet.cell(row=row, column=5, value=trade_value)
            trade_val_cell.number_format = "#,##0.00"

            pct_tna_cell = sheet.cell(row=row, column=6, value=pct_tna)
            pct_tna_cell.number_format = "0.00%"

            pct_assets_cell = sheet.cell(row=row, column=7, value=pct_assets)
            pct_assets_cell.number_format = "0.00%"

            ex_ante_cell = sheet.cell(row=row, column=8, value=ex_ante)
            ex_ante_cell.number_format = "#,##0.00"

            ex_post_cell = sheet.cell(row=row, column=9, value=ex_post)
            ex_post_cell.number_format = "#,##0.00"

            delta_cell = sheet.cell(row=row, column=10, value=market_delta)
            delta_cell.number_format = "#,##0.00"

            trade_vs_ante_cell = sheet.cell(row=row, column=11, value=trade_vs_ante)
            trade_vs_ante_cell.number_format = "0.00%"

            post_vs_ante_cell = sheet.cell(row=row, column=12, value=post_vs_ante)
            post_vs_ante_cell.number_format = "0.00%"

            net_trade_value_cell = sheet.cell(row=row, column=13, value=net_value)
            net_trade_value_cell.number_format = "#,##0.00"

            net_buy_qty_cell = sheet.cell(row=row, column=14, value=buy_qty)
            net_buy_qty_cell.number_format = "#,##0.00"

            net_sell_qty_cell = sheet.cell(row=row, column=15, value=sell_qty)
            net_sell_qty_cell.number_format = "#,##0.00"

            net_buy_value_cell = sheet.cell(row=row, column=16, value=buy_val)
            net_buy_value_cell.number_format = "#,##0.00"

            net_sell_value_cell = sheet.cell(row=row, column=17, value=sell_val)
            net_sell_value_cell.number_format = "#,##0.00"

            row += 1

    if not summary_rows_written:
        sheet.cell(row=header_row + 1, column=1, value="No trading activity detected.")

    for col in range(1, len(summary_headers) + 1):
        column = get_column_letter(col)
        width = sheet.column_dimensions[column].width or 0
        desired = 22 if col <= 5 else 18
        if col in (1, 2):
            desired = 24 if col == 1 else 20
        if width < desired:
            sheet.column_dimensions[column].width = desired

    # Trade Activity Details section
    row = max(row + 2, header_row + 3)
    sheet.cell(row=row, column=1, value="Trade Activity Details (Top 5 by Market Value)").font = Font(size=12, bold=True)
    row += 1

    detail_headers = [
        "Fund",
        "Asset Class",
        "Direction",
        "Ticker",
        "Quantity",
        "Market Value",
    ]

    for col, header in enumerate(detail_headers, start=1):
        cell = sheet.cell(row=row, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    row += 1
    detail_rows_written = False

    for fund_name, fund_data in sorted(data.get("funds", {}).items()):
        trade_info = fund_data.get("trade_info", {}) or {}
        sections = _build_asset_sections(trade_info)

        fund_rows_written = False

        for section in sections:
            activity_details = section.get("activity", {}) or {}
            if not isinstance(activity_details, Mapping):
                continue

            asset_label = section.get("label", section.get("asset_key", "")).title()

            # Only show top 5 trades for equity and options
            if section.get("asset_key") in ("equity", "options"):
                buys_list = _get_top_trades(activity_details.get("buys", []), n=5)
                sells_list = _get_top_trades(activity_details.get("sells", []), n=5)
            else:
                # Show all trades for treasury or other assets
                buys_list = activity_details.get("buys", []) or []
                sells_list = activity_details.get("sells", []) or []

            net_info = section.get("net", {}) or {}
            buy_qty = float(net_info.get("buy_quantity", 0.0) or 0.0)
            sell_qty = float(net_info.get("sell_quantity", 0.0) or 0.0)
            buy_val = float(net_info.get("buy_value", 0.0) or 0.0)
            sell_val = float(net_info.get("sell_value", 0.0) or 0.0)

            if not buys_list and not sells_list and all(
                    abs(val) == 0.0 for val in (buy_qty, sell_qty, buy_val, sell_val)
            ):
                continue

            def _write_row(direction: str, ticker: str, quantity: float, value: float) -> None:
                nonlocal row, fund_rows_written, detail_rows_written
                sheet.cell(row=row, column=1, value=fund_name if not fund_rows_written else "")
                sheet.cell(row=row, column=2, value=asset_label)
                sheet.cell(row=row, column=3, value=direction)
                sheet.cell(row=row, column=4, value=ticker)
                qty_cell = sheet.cell(row=row, column=5, value=quantity)
                qty_cell.number_format = "#,##0.00"
                val_cell = sheet.cell(row=row, column=6, value=value)
                val_cell.number_format = "#,##0.00"
                row += 1
                detail_rows_written = True
                fund_rows_written = True

            # Always show net totals
            if abs(buy_qty) > 0.0 or abs(buy_val) > 0.0:
                _write_row("Net Buy", "", buy_qty, buy_val)

            if abs(sell_qty) > 0.0 or abs(sell_val) > 0.0:
                _write_row("Net Sell", "", sell_qty, sell_val)

            # Show individual trades (limited to top 5 for equity/options)
            for trade in buys_list:
                quantity = float(trade.get("quantity", 0.0) or 0.0)
                value = float(trade.get("market_value", 0.0) or 0.0)
                if quantity == 0.0 and value == 0.0:
                    continue
                ticker = str(trade.get("eqyticker", ""))
                _write_row("Buy", ticker, quantity, value)

            for trade in sells_list:
                quantity = float(trade.get("quantity", 0.0) or 0.0)
                value = float(trade.get("market_value", 0.0) or 0.0)
                if quantity == 0.0 and value == 0.0:
                    continue
                ticker = str(trade.get("ticker", ""))
                _write_row("Sell", ticker, quantity, value)

            if fund_rows_written:
                row += 1

    if not detail_rows_written:
        sheet.cell(row=row, column=1, value="No trade details available.")

    for col in range(1, len(detail_headers) + 1):
        column = get_column_letter(col)
        width = sheet.column_dimensions[column].width or 0
        desired = 18 if col >= 4 else 20
        if width < desired:
            sheet.column_dimensions[column].width = desired

def _populate_compliance_detail_sheet(sheet, data: Mapping[str, Any]) -> None:
    headers = [
        "Fund Name",
        "Compliance Check",
        "Status Before",
        "Status After",
        "Violations Before",
        "Violations After",
        "Changed",
    ]

    for col, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

    row = 2
    for fund_name, fund_data in sorted(data.get("funds", {}).items()):
        for check_name, check_data in sorted(fund_data.get("checks", {}).items()):
            sheet.cell(row=row, column=1, value=fund_name)
            sheet.cell(row=row, column=2, value=check_name)
            sheet.cell(row=row, column=3, value=check_data.get("status_before", "UNKNOWN"))
            sheet.cell(row=row, column=4, value=check_data.get("status_after", "UNKNOWN"))
            sheet.cell(row=row, column=5, value=check_data.get("violations_before", 0))
            violations_after = int(check_data.get("violations_after", 0) or 0)
            sheet.cell(row=row, column=6, value=violations_after)
            changed = bool(check_data.get("changed"))
            sheet.cell(row=row, column=7, value="YES" if changed else "NO")

            if violations_after > 0:
                fill = PatternFill(
                    start_color="FF9999", end_color="FF9999", fill_type="solid"
                )
                for col in range(1, len(headers) + 1):
                    sheet.cell(row=row, column=col).fill = fill
            elif changed:
                for col in range(1, len(headers) + 1):
                    sheet.cell(row=row, column=col).fill = PatternFill(
                        start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"
                    )

            row += 1

    for col in range(1, len(headers) + 1):
        sheet.column_dimensions[get_column_letter(col)].width = 22

def _create_compliance_details_sheet(workbook: Workbook, data: Mapping[str, Any]) -> None:
    """Create the compliance details sheet (mirrors detailed comparison)."""

    sheet = workbook.create_sheet("Compliance Details")
    _populate_compliance_detail_sheet(sheet, data)




class TradingCompliancePDF(FPDF):
    """Custom PDF class for trading compliance reports."""

    def __init__(self, date: str) -> None:
        super().__init__()
        self.date = date
        self.set_auto_page_break(auto=True, margin=15)

    def header(self) -> None:  # pragma: no cover - provided by FPDF
        self.set_font("Arial", "B", 16)
        self.cell(0, 10, "Trading Compliance Analysis", 0, 1, "C")
        self.set_font("Arial", "", 10)
        self.cell(0, 8, f"Ex-Ante vs Ex-Post Comparison - {self.date}", 0, 1, "C")
        self.ln(5)

    def footer(self) -> None:  # pragma: no cover - provided by FPDF
        self.set_y(-15)
        self.set_font("Arial", "I", 8)
        self.cell(0, 10, f"Page {self.page_no()}", 0, 0, "C")


def generate_trading_pdf_report(comparison_data: Mapping[str, Any], output_path: str) -> None:
    """Create a PDF report for the trading compliance comparison."""

    try:
        pdf = TradingCompliancePDF(date=str(comparison_data.get("date", "")))
        pdf.add_page(orientation="L")
        _add_compliance_overview(pdf, comparison_data)

        _add_summary_metrics_section(pdf, comparison_data)

        pdf.add_page(orientation="L")
        _add_detailed_comparison(pdf, comparison_data)

        pdf.add_page(orientation="L")
        _add_trade_activity(pdf, comparison_data)

        Path(output_path).parent.mkdir(parents=True, exist_ok=True)
        pdf.output(output_path)
        logger.info("Trading PDF report saved to: %s", output_path)
    except Exception as exc:  # pragma: no cover - defensive
        logger.error("Error generating trading PDF report: %s", exc)
        raise


def _add_compliance_overview(pdf: FPDF, data: Mapping[str, Any]) -> None:
    summary = data.get("summary", {}) or {}

    pdf.set_font("Arial", "B", 12)
    pdf.cell(0, 8, "Compliance Status Changes", 0, 1, "L")
    pdf.ln(1)

    pdf.set_font("Arial", "", 9)
    pdf.cell(
        0,
        6,
        f"Total Funds Analyzed: {summary.get('total_funds_analyzed', 0)}",
        0,
        1,
        "L",
    )
    pdf.cell(
        0,
        6,
        f"Total Funds Traded: {summary.get('total_funds_traded', 0)}",
        0,
        1,
        "L",
    )
    pdf.ln(2)

    _render_compliance_changes_table(pdf, data)
    pdf.ln(3)


def _render_compliance_changes_table(pdf: FPDF, data: Mapping[str, Any]) -> None:
    col_widths = [55, 40, 25, 25, 25]
    headers = ["Fund", "Status Change", "Viol. Before", "Viol. After", "Net Change"]

    def _draw_header_row(title: Optional[str] = None) -> None:
        if title:
            pdf.set_font("Arial", "B", 12)
            pdf.cell(0, 8, title, 0, 1, "L")
            pdf.ln(1)
        pdf.set_font("Arial", "B", 8)
        pdf.set_fill_color(200, 200, 200)
        for width, header in zip(col_widths, headers):
            pdf.cell(width, 6, header, 1, 0, "C", True)
        pdf.ln()
        pdf.set_font("Arial", "", 8)

    _draw_header_row()

    total_before = 0
    total_after = 0
    total_net = 0

    for fund_name, fund_data in sorted(data.get("funds", {}).items()):
        if pdf.get_y() > 260:
            pdf.add_page()
            _draw_header_row("Compliance Status Changes (cont.)")

        display_name = fund_name if len(fund_name) <= 26 else f"{fund_name[:23]}..."
        status_change = str(fund_data.get("status_change", "UNCHANGED"))
        status_upper = status_change.upper()
        viol_before = int(fund_data.get("violations_before", 0) or 0)
        viol_after = int(fund_data.get("violations_after", 0) or 0)
        net_change = viol_after - viol_before

        total_before += viol_before
        total_after += viol_after
        total_net += net_change

        if status_upper == "INTO_COMPLIANCE":
            pdf.set_fill_color(204, 255, 204)
            fill = True
        elif status_upper == "OUT_OF_COMPLIANCE":
            pdf.set_fill_color(255, 204, 153)
            fill = True
        elif status_upper == "UNCHANGED" and viol_after > 0:
            pdf.set_fill_color(255, 153, 153)
            fill = True
        else:
            fill = False

        pdf.cell(col_widths[0], 6, display_name, 1, 0, "L", fill)
        pdf.cell(col_widths[1], 6, status_change, 1, 0, "C", fill)
        pdf.cell(col_widths[2], 6, str(viol_before), 1, 0, "C", fill)
        pdf.cell(col_widths[3], 6, str(viol_after), 1, 0, "C", fill)
        pdf.cell(col_widths[4], 6, str(net_change), 1, 1, "C", fill)

    if pdf.get_y() > 260:
        pdf.add_page()
        _draw_header_row("Compliance Status Changes (cont.)")

    pdf.set_font("Arial", "B", 8)
    pdf.set_fill_color(220, 220, 220)
    pdf.cell(col_widths[0], 6, "Totals", 1, 0, "L", True)
    pdf.cell(col_widths[1], 6, "", 1, 0, "C", True)
    pdf.cell(col_widths[2], 6, str(total_before), 1, 0, "C", True)
    pdf.cell(col_widths[3], 6, str(total_after), 1, 0, "C", True)
    pdf.cell(col_widths[4], 6, str(total_net), 1, 1, "C", True)
    pdf.set_font("Arial", "", 8)

def _get_top_trades(trades_list: list, n: int = 5) -> list:
    """Get top N trades sorted by market value."""
    if not trades_list:
        return []
    # Sort by market value (descending) and take top N
    sorted_trades = sorted(trades_list, key=lambda x: abs(float(x.get("market_value", 0))), reverse=True)
    return sorted_trades[:n]

def _add_trade_activity(pdf: FPDF, data: Mapping[str, Any]) -> None:
    pdf.set_font("Arial", "B", 14)
    pdf.cell(0, 10, "Trade Activity", 0, 1, "L")
    pdf.ln(3)

    funds_payload = data.get("funds", {}) or {}

    summary_rows: list[dict[str, Any]] = []
    detail_rows: list[tuple[str, str, str, str, float, float]] = []

    for fund_name, fund_data in sorted(funds_payload.items()):
        trade_info = fund_data.get("trade_info", {}) or {}
        sections = _build_asset_sections(trade_info)

        total_net_assets = float(trade_info.get("total_net_assets", 0.0) or 0.0)
        total_assets = float(trade_info.get("total_assets", 0.0) or 0.0)

        for section in sections:
            summary = section.get("summary", {}) or {}
            activity_details = section.get("activity", {}) or {}
            net_info = section.get("net", {}) or {}

            trade_value = float(summary.get("trade_value", 0.0) or 0.0)
            market_delta = float(summary.get("market_value_delta", 0.0) or 0.0)
            pct_tna = float(summary.get("pct_of_tna", 0.0) or 0.0)
            pct_assets = float(summary.get("pct_of_total_assets", 0.0) or 0.0)
            ex_ante = float(summary.get("ex_ante_market_value", 0.0) or 0.0)
            ex_post = float(summary.get("ex_post_market_value", 0.0) or 0.0)

            buy_qty = float(net_info.get("buy_quantity", net_info.get("buys", 0.0)) or 0.0)
            sell_qty = float(net_info.get("sell_quantity", net_info.get("sells", 0.0)) or 0.0)
            buy_val = float(net_info.get("buy_value", 0.0) or 0.0)
            sell_val = float(net_info.get("sell_value", 0.0) or 0.0)

            if trade_value == 0.0:
                trade_value = abs(buy_val) + abs(sell_val)

            raw_net_value = net_info.get("net_value")
            if raw_net_value in (None, ""):
                net_value = buy_val - sell_val
            else:
                try:
                    net_value = float(raw_net_value)
                except (TypeError, ValueError):
                    net_value = buy_val - sell_val

            buys_list = activity_details.get("buys") if isinstance(activity_details, Mapping) else []
            sells_list = activity_details.get("sells") if isinstance(activity_details, Mapping) else []

            if section.get("asset_key") in ("equity", "options"):
                buys_list = _get_top_trades(buys_list or [], n=5)
                sells_list = _get_top_trades(sells_list or [], n=5)

            has_trade_lists = bool(buys_list or sells_list)

            if not has_trade_lists and not any(
                abs(val) > 0
                for val in (
                    trade_value,
                    market_delta,
                    net_value,
                    buy_qty,
                    sell_qty,
                    buy_val,
                    sell_val,
                )
            ):
                continue

            asset_label = section.get("label", section.get("asset_key", "")).title()

            summary_rows.append(
                {
                    "fund": fund_name,
                    "asset": asset_label,
                    "total_net_assets": total_net_assets,
                    "total_assets": total_assets,
                    "trade_value": trade_value,
                    "pct_tna": pct_tna,
                    "pct_assets": pct_assets,
                    "ex_ante": ex_ante,
                    "ex_post": ex_post,
                    "delta": market_delta,
                    "net_value": net_value,
                    "buy_value": buy_val,
                    "sell_value": sell_val,
                }
            )

            if abs(buy_qty) > 0.0 or abs(buy_val) > 0.0:
                detail_rows.append((fund_name, asset_label, "Net Buy", "", buy_qty, buy_val))
            if abs(sell_qty) > 0.0 or abs(sell_val) > 0.0:
                detail_rows.append((fund_name, asset_label, "Net Sell", "", sell_qty, sell_val))

            for trade in buys_list or []:
                quantity = float(trade.get("quantity", 0.0) or 0.0)
                value = float(trade.get("market_value", 0.0) or 0.0)
                if not quantity and not value:
                    continue
                detail_rows.append(
                    (
                        fund_name,
                        asset_label,
                        "Buy",
                        str(trade.get("ticker", "")),
                        quantity,
                        value,
                    )
                )

            for trade in sells_list or []:
                quantity = float(trade.get("quantity", 0.0) or 0.0)
                value = float(trade.get("market_value", 0.0) or 0.0)
                if not quantity and not value:
                    continue
                detail_rows.append(
                    (
                        fund_name,
                        asset_label,
                        "Sell",
                        str(trade.get("ticker", "")),
                        quantity,
                        value,
                    )
                )

    if not summary_rows and not detail_rows:
        pdf.set_font("Arial", "", 9)
        pdf.cell(0, 6, "No trading activity detected.", 0, 1, "L")
        return

    usable_width = pdf.w - pdf.l_margin - pdf.r_margin
    summary_headers = [
        "Fund",
        "Asset Class",
        "Total Net Assets",
        "Total Assets",
        "Trade Value",
        "% of TNA",
        "% of Assets",
        "Ex-Ante MV",
        "Ex-Post MV",
        "Market Value Delta",
        "Net Trade Value",
        "Net Buy Value",
        "Net Sell Value",
    ]
    summary_rel_widths = [1.1, 1.0, 1.15, 1.1, 1.1, 0.85, 0.85, 1.1, 1.1, 1.1, 1.0, 1.0, 1.0]
    total_weight = sum(summary_rel_widths)
    summary_widths = [usable_width * w / total_weight for w in summary_rel_widths]

    pdf.set_font("Arial", "B", 10)
    pdf.cell(0, 6, "Trade Activity Summary", 0, 1, "L")
    pdf.ln(1)

    pdf.set_font("Arial", "B", 6)
    pdf.set_fill_color(200, 200, 200)
    for width, header in zip(summary_widths, summary_headers):
        pdf.cell(width, 5.5, header, 1, 0, "C", True)
    pdf.ln(5.5)

    pdf.set_font("Arial", "", 6)
    page_limit = pdf.h - pdf.b_margin - 10
    row_height = 5.0
    last_fund = None

    for row in sorted(summary_rows, key=lambda item: (item["fund"], item["asset"])):
        if pdf.get_y() + row_height > page_limit:
            pdf.add_page(orientation=getattr(pdf, "cur_orientation", "L"))
            pdf.set_font("Arial", "B", 6)
            pdf.set_fill_color(200, 200, 200)
            for width, header in zip(summary_widths, summary_headers):
                pdf.cell(width, 5.5, header, 1, 0, "C", True)
            pdf.ln(5.5)
            pdf.set_font("Arial", "", 6)

        fund_label = row["fund"] if row["fund"] != last_fund else ""
        if fund_label:
            pdf.set_font("Arial", "B", 6)
        pdf.cell(summary_widths[0], row_height, fund_label, 1, 0, "L")
        if fund_label:
            pdf.set_font("Arial", "", 6)
        pdf.cell(summary_widths[1], row_height, row["asset"], 1, 0, "L")

        pdf.cell(summary_widths[2], row_height, format_number(row["total_net_assets"], digits=2), 1, 0, "R")
        pdf.cell(summary_widths[3], row_height, format_number(row["total_assets"], digits=2), 1, 0, "R")
        pdf.cell(summary_widths[4], row_height, format_number(row["trade_value"], digits=2), 1, 0, "R")
        pdf.cell(summary_widths[5], row_height, _format_percent(row["pct_tna"]), 1, 0, "R")
        pdf.cell(summary_widths[6], row_height, _format_percent(row["pct_assets"]), 1, 0, "R")
        pdf.cell(summary_widths[7], row_height, format_number(row["ex_ante"], digits=2), 1, 0, "R")
        pdf.cell(summary_widths[8], row_height, format_number(row["ex_post"], digits=2), 1, 0, "R")
        pdf.cell(summary_widths[9], row_height, format_number(row["delta"], digits=2), 1, 0, "R")
        pdf.cell(summary_widths[10], row_height, format_number(row["net_value"], digits=2), 1, 0, "R")
        pdf.cell(summary_widths[11], row_height, format_number(row["buy_value"], digits=2), 1, 0, "R")
        pdf.cell(summary_widths[12], row_height, format_number(row["sell_value"], digits=2), 1, 1, "R")

        last_fund = row["fund"]

    pdf.ln(3)

    if detail_rows:
        detail_headers = ["Fund", "Asset Class", "Direction", "Ticker", "Quantity", "Market Value"]
        detail_rel_widths = [1.2, 1.0, 0.9, 1.0, 0.9, 1.1]
        detail_total = sum(detail_rel_widths)
        detail_widths = [usable_width * w / detail_total for w in detail_rel_widths]

        pdf.set_font("Arial", "B", 10)
        pdf.cell(0, 6, "Trade Activity Details", 0, 1, "L")
        pdf.ln(1)

        pdf.set_font("Arial", "B", 6)
        pdf.set_fill_color(200, 200, 200)
        for width, header in zip(detail_widths, detail_headers):
            pdf.cell(width, 5.5, header, 1, 0, "C", True)
        pdf.ln(5.5)

        pdf.set_font("Arial", "", 6)
        last_fund = None
        row_height = 5.0

        for fund_name, asset, direction, ticker, quantity, value in detail_rows:
            if pdf.get_y() + row_height > page_limit:
                pdf.add_page(orientation=getattr(pdf, "cur_orientation", "L"))
                pdf.set_font("Arial", "B", 6)
                pdf.set_fill_color(200, 200, 200)
                for width, header in zip(detail_widths, detail_headers):
                    pdf.cell(width, 5.5, header, 1, 0, "C", True)
                pdf.ln(5.5)
                pdf.set_font("Arial", "", 6)

            fund_label = fund_name if fund_name != last_fund else ""
            if fund_label:
                pdf.set_font("Arial", "B", 6)
            pdf.cell(detail_widths[0], row_height, fund_label, 1, 0, "L")
            if fund_label:
                pdf.set_font("Arial", "", 6)

            pdf.cell(detail_widths[1], row_height, asset, 1, 0, "L")
            pdf.cell(detail_widths[2], row_height, direction, 1, 0, "L")
            pdf.cell(detail_widths[3], row_height, ticker, 1, 0, "L")
            pdf.cell(detail_widths[4], row_height, format_number(quantity, digits=2), 1, 0, "R")
            pdf.cell(detail_widths[5], row_height, format_number(value, digits=2), 1, 1, "R")

            last_fund = fund_name


def _add_summary_metrics_section(pdf: FPDF, data: Mapping[str, Any]) -> None:
    metric_labels = {
        "cash_value": "Cash Value",
        "equity_market_value": "Equity Market Value",
        "option_market_value": "Option Market Value",
        "option_delta_adjusted_notional": "Option Delta Adjusted Notional",
        "treasury": "Treasury Market Value",
        "total_assets": "Total Assets",
        "total_net_assets": "Total Net Assets",
    }

    fund_entries: list[tuple[str, Dict[str, Any], Dict[str, Any]]] = []
    for fund_name, fund_payload in sorted(data.get("funds", {}).items()):
        summary_metrics = fund_payload.get("summary_metrics", {}) or {}
        ex_ante = dict(summary_metrics.get("ex_ante", {}) or {})
        ex_post = dict(summary_metrics.get("ex_post", {}) or {})
        if ex_ante or ex_post:
            fund_entries.append((fund_name, ex_ante, ex_post))

    if not fund_entries:
        return

    table_rows: list[tuple[str, list[tuple[float, float, float]]]] = []
    for metric_key, label in metric_labels.items():
        metric_values: list[tuple[float, float, float]] = []
        metric_present = False
        for _, ex_ante, ex_post in fund_entries:
            ante_value = float(ex_ante.get(metric_key, 0.0) or 0.0)
            post_value = float(ex_post.get(metric_key, 0.0) or 0.0)
            if metric_key in ex_ante or metric_key in ex_post:
                metric_present = True
            metric_values.append((ante_value, post_value, post_value - ante_value))
        if metric_present:
            table_rows.append((label, metric_values))

    if not table_rows:
        return

    usable_width = pdf.w - pdf.l_margin - pdf.r_margin
    metric_width = min(60.0, max(40.0, usable_width * 0.22))
    remaining_width = max(usable_width - metric_width, 60.0)
    per_value_width = remaining_width / (len(fund_entries) * 3)

    header_height = 6
    row_height = 5.5
    page_limit = pdf.h - pdf.b_margin - 5

    def _draw_header(title: str) -> None:
        pdf.set_font("Arial", "B", 12)
        pdf.cell(0, 8, title, 0, 1, "L")
        pdf.ln(1)

        start_x = pdf.l_margin
        start_y = pdf.get_y()

        pdf.set_font("Arial", "B", 7)
        pdf.set_fill_color(200, 200, 200)

        metric_height = header_height * 2
        pdf.cell(metric_width, metric_height, "Metric", 1, 0, "C", True)
        for fund_name, _, _ in fund_entries:
            display = fund_name if len(fund_name) <= 24 else f"{fund_name[:21]}..."
            pdf.cell(per_value_width * 3, header_height, display, 1, 0, "C", True)

        pdf.ln(header_height)
        pdf.set_xy(start_x + metric_width, start_y + header_height)
        for _ in fund_entries:
            for sub in ("Ex-Ante", "Ex-Post", "Delta"):
                pdf.cell(per_value_width, header_height, sub, 1, 0, "C", True)

        pdf.ln(header_height)
        pdf.set_y(start_y + metric_height)
        pdf.set_font("Arial", "", 7)

    if pdf.get_y() > page_limit:
        pdf.add_page(orientation=getattr(pdf, "cur_orientation", "L"))

    _draw_header("Fund Summary Metrics")

    for label, values in table_rows:
        if pdf.get_y() + row_height > page_limit:
            pdf.add_page(orientation=getattr(pdf, "cur_orientation", "L"))
            _draw_header("Fund Summary Metrics (cont.)")

        pdf.set_x(pdf.l_margin)
        pdf.set_font("Arial", "B", 7)
        pdf.cell(metric_width, row_height, label, 1, 0, "L")
        pdf.set_font("Arial", "", 7)

        for ante_value, post_value, delta in values:
            pdf.cell(per_value_width, row_height, format_number(ante_value, digits=2), 1, 0, "R")
            pdf.cell(per_value_width, row_height, format_number(post_value, digits=2), 1, 0, "R")
            pdf.cell(per_value_width, row_height, format_number(delta, digits=2), 1, 0, "R")

        pdf.ln(row_height)

    pdf.ln(3)

def _add_detailed_comparison(pdf: FPDF, data: Mapping[str, Any]) -> None:
    funds = []
    for fund_name, fund_payload in sorted(data.get("funds", {}).items()):
        checks = fund_payload.get("checks", {}) or {}
        if checks:
            funds.append((fund_name, checks))

    if not funds:
        return

    check_names: list[str] = []
    seen_checks: set[str] = set()
    for _, checks in funds:
        for check_name in checks.keys():
            if check_name not in seen_checks:
                seen_checks.add(check_name)
                check_names.append(check_name)
    check_names.sort()

    if not check_names:
        return

    table_rows: list[tuple[str, list[tuple[str, str, str, str]]]] = []
    for check_name in check_names:
        row_values: list[tuple[str, str, str, str]] = []
        row_present = False
        for _, checks in funds:
            check_data = checks.get(check_name) or {}
            if check_data:
                row_present = True
            before_status = str(check_data.get("status_before", "")) if check_data else ""
            after_status = str(check_data.get("status_after", "")) if check_data else ""
            changed_flag = "YES" if bool(check_data.get("changed")) else ("" if not check_data else "NO")
            row_values.append((before_status, after_status, changed_flag, after_status.upper()))
        if row_present:
            table_rows.append((check_name, row_values))

    if not table_rows:
        return

    header_height = 6
    row_height = 6
    page_limit = pdf.h - pdf.b_margin - 5

    fund_order = [fund_name for fund_name, _ in funds]
    fund_index = {name: idx for idx, name in enumerate(fund_order)}
    fund_chunks = [funds[i:i + 5] for i in range(0, len(funds), 5)]

    def _render_header(title: str, fund_subset: list[tuple[str, Any]], per_value_width: float, metric_width: float) -> None:
        pdf.set_font("Arial", "B", 14)
        pdf.cell(0, 10, title, 0, 1, "L")
        pdf.ln(1)

        start_x = pdf.l_margin
        start_y = pdf.get_y()

        pdf.set_font("Arial", "B", 7)
        pdf.set_fill_color(200, 200, 200)

        metric_height = header_height * 2
        pdf.cell(metric_width, metric_height, "Compliance Check", 1, 0, "C", True)
        for fund_name, _ in fund_subset:
            display = fund_name if len(fund_name) <= 24 else f"{fund_name[:21]}..."
            pdf.cell(per_value_width * 3, header_height, display, 1, 0, "C", True)

        pdf.ln(header_height)
        pdf.set_xy(start_x + metric_width, start_y + header_height)
        for _ in fund_subset:
            for sub in ("Before", "After", "Changed"):
                pdf.cell(per_value_width, header_height, sub, 1, 0, "C", True)

        pdf.ln(header_height)
        pdf.set_y(start_y + metric_height)
        pdf.set_font("Arial", "", 7)

    for chunk_index, fund_subset in enumerate(fund_chunks):
        chunk_start = chunk_index * 5
        usable_width = pdf.w - pdf.l_margin - pdf.r_margin
        metric_width = min(70.0, max(45.0, usable_width * 0.25))
        remaining_width = max(usable_width - metric_width, 60.0)
        per_value_width = remaining_width / (len(fund_subset) * 3)

        title = "Detailed Comparison by Compliance Check"
        if chunk_index:
            title = f"{title} (Group {chunk_index + 1})"

        _render_header(title, fund_subset, per_value_width, metric_width)

        for check_name, values in table_rows:
            if pdf.get_y() + row_height > page_limit:
                pdf.add_page(orientation=getattr(pdf, "cur_orientation", "L"))
                _render_header(f"{title} (cont.)", fund_subset, per_value_width, metric_width)

            pdf.set_x(pdf.l_margin)
            display_check = check_name if len(check_name) <= 40 else f"{check_name[:37]}..."
            pdf.set_font("Arial", "B", 7)
            pdf.cell(metric_width, row_height, display_check, 1, 0, "L")
            pdf.set_font("Arial", "", 7)

            subset_indices = [
                fund_index.get(fund_name, chunk_start + offset)
                for offset, (fund_name, _) in enumerate(fund_subset)
            ]
            subset_values = [
                values[idx] if idx is not None and idx < len(values) else ("", "", "", "")
                for idx in subset_indices
            ]

            for before_status, after_status, changed_flag, after_upper in subset_values:
                pdf.cell(per_value_width, row_height, before_status, 1, 0, "C")
                if after_upper == "FAIL":
                    pdf.set_fill_color(255, 204, 204)
                    pdf.cell(per_value_width, row_height, after_status, 1, 0, "C", True)
                    pdf.set_fill_color(255, 255, 255)
                else:
                    pdf.cell(per_value_width, row_height, after_status, 1, 0, "C")

                changed_display = changed_flag or ""
                if changed_display.upper() == "YES":
                    pdf.set_fill_color(255, 255, 204)
                    pdf.cell(per_value_width, row_height, changed_display, 1, 0, "C", True)
                    pdf.set_fill_color(255, 255, 255)
                else:
                    pdf.cell(per_value_width, row_height, changed_display, 1, 0, "C")

            pdf.ln(row_height)

        pdf.ln(3)

def generate_trading_compliance_reports(
    comparison_data: Mapping[str, Any],
    report_date: Any,
    output_dir: str,
    *,
    file_name_prefix: str = "trading_compliance_results",
    create_pdf: bool = True,
) -> GeneratedTradingComplianceReport:
    """Generate Excel and PDF outputs for trading compliance comparisons."""

    if not comparison_data:
        return GeneratedTradingComplianceReport(None, None)

    date_str = normalize_report_date(report_date)
    output_path = Path(output_dir)
    output_path.mkdir(parents=True, exist_ok=True)

    excel_path = output_path / f"{file_name_prefix}_{date_str}.xlsx"
    pdf_path = output_path / f"{file_name_prefix}_{date_str}.pdf"

    generate_trading_excel_report(comparison_data, str(excel_path))

    pdf_result: Optional[str] = None
    if create_pdf:
        generate_trading_pdf_report(comparison_data, str(pdf_path))
        pdf_result = str(pdf_path)

    return GeneratedTradingComplianceReport(str(excel_path), pdf_result)


__all__ = [
    "GeneratedTradingComplianceReport",
    "generate_trading_compliance_reports",
    "generate_trading_excel_report",
    "generate_trading_pdf_report",
]