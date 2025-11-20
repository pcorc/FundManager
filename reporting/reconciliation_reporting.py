"""Holdings reconciliation reporting utilities."""
from __future__ import annotations
from openpyxl.styles import Alignment, Font, PatternFill
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any, Callable, Dict, Iterable, Mapping, Optional, Sequence, Tuple

import numpy as np
import pandas as pd

from config.fund_definitions import FUND_DEFINITIONS, INDEX_FLEX_FUNDS
from reporting.report_utils import normalize_reconciliation_payload, normalize_report_date
from utilities.reconciliation_utils import split_flex_price_frames


@dataclass
class GeneratedReconciliationReport:
    """Container for holdings reconciliation artefact locations."""
    excel_path: Optional[str]

@dataclass(frozen=True)
class ReconDescriptor:
    """Declarative definition of how to flatten reconciliation payloads."""

    holdings_key: str | None = None
    holdings_ticker: str | None = None
    holdings_builder: str | None = None
    require_holdings: bool = True
    price_keys: Tuple[str | None, str | None] = ("price_discrepancies_T", "price_discrepancies_T1")
    price_ticker: str | None = None
    price_cust_col: str = "price_cust"
    price_transform: str | None = None
    extra_callbacks: Sequence[str] = ()


RECON_DESCRIPTOR_REGISTRY: Dict[str, ReconDescriptor] = {
    "custodian_equity": ReconDescriptor(
        holdings_key="final_recon",
        holdings_ticker="eqyticker",
        price_ticker="eqyticker",
        extra_callbacks=("_append_holdings_breakdowns",),
    ),
    "custodian_equity_t1": ReconDescriptor(
        holdings_key="final_recon",
        holdings_ticker="eqyticker",
        price_keys=(None, None),
        extra_callbacks=("_append_holdings_breakdowns",),
    ),
    "custodian_option": ReconDescriptor(
        holdings_key="final_recon",
        holdings_ticker="optticker",
        price_ticker="optticker",
        price_transform="_select_standard_option_prices",
        extra_callbacks=("_append_option_breakdowns","_append_holdings_breakdowns"),
    ),
    "custodian_option_t1": ReconDescriptor(
        holdings_key="final_recon",
        holdings_ticker="optticker",
        price_keys=(None, None),
        extra_callbacks=("_append_option_t1_breakdowns","_append_holdings_breakdowns"),
    ),
    "index_equity": ReconDescriptor(
        holdings_key="holdings_discrepancies",
        holdings_ticker="eqyticker",
        holdings_builder="_prepare_index_holdings_df",
        price_ticker="eqyticker",
        price_cust_col="price_index",
    ),
    # "sg_option": ReconDescriptor(
    #     holdings_key="final_recon",
    #     holdings_ticker="optticker",
    #     price_keys=("price_discrepancies", None),
    #     price_ticker="optticker",
    # ),
    # "sg_equity": ReconDescriptor(
    #     holdings_key="final_recon",
    #     holdings_ticker="equity_ticker",
    #     price_keys=("price_discrepancies", None),
    #     price_ticker="equity_ticker",
    # ),
    "custodian_treasury": ReconDescriptor(
        holdings_key="final_recon",
        holdings_ticker="cusip",
        price_keys=("price_discrepancies_T", None),
        price_ticker="cusip",
        require_holdings=False,
        extra_callbacks=("_append_holdings_breakdowns",),
    ),
    "custodian_treasury_t1": ReconDescriptor(
        holdings_key="final_recon",
        holdings_ticker="cusip",
        price_keys=(None, "price_discrepancies_T1"),
        price_ticker="cusip",
        require_holdings=False,
        extra_callbacks=("_append_holdings_breakdowns",),
    ),
}


class ReconResult:
    """Standardized container used when flattening reconciliation payloads."""

    def __init__(
        self,
        fund: str,
        date_str: str,
        recon_type: str,
        holdings_df: pd.DataFrame | None = None,
        price_df_t: pd.DataFrame | None = None,
        price_df_t1: pd.DataFrame | None = None,
    ) -> None:
        self.fund = fund
        self.date_str = date_str
        self.recon_type = recon_type
        self.holdings_df = holdings_df
        self.price_df_t = price_df_t
        self.price_df_t1 = price_df_t1

    def to_flat_dict(self) -> Dict[Tuple[str, str, str], pd.DataFrame]:
        flat: Dict[Tuple[str, str, str], pd.DataFrame] = {}
        if self.holdings_df is not None and not self.holdings_df.empty:
            flat[(self.fund, self.date_str, self.recon_type)] = self.holdings_df
        if self.price_df_t is not None and not self.price_df_t.empty:
            flat[(self.fund, self.date_str, f"{self.recon_type}_price_T")] = self.price_df_t
        if self.price_df_t1 is not None and not self.price_df_t1.empty:
            flat[(self.fund, self.date_str, f"{self.recon_type}_price_T-1")] = self.price_df_t1
        return flat

    def has_data(self) -> bool:
        return any(
            [
                self.holdings_df is not None and not self.holdings_df.empty,
                self.price_df_t is not None and not self.price_df_t.empty,
                self.price_df_t1 is not None and not self.price_df_t1.empty,
            ]
        )


class ReconciliationReport:
    """Generate the enhanced Excel holdings reconciliation report."""
    SUMMARY_MAPPING: Dict[str, str] = {
        "custodian_equity": "custodian_equity_holdings_breaks_T",
        "custodian_equity_price_T": "custodian_equity_price_breaks_T",
        "custodian_option": "custodian_option_holdings_breaks_T",
        "custodian_option_price_T": "custodian_option_price_breaks_T",
        "custodian_treasury": "custodian_treasury_holdings_breaks_T",
        "custodian_treasury_price_T": "custodian_treasury_price_breaks_T",
        "index_equity": "index_equity_holdings_breaks",
        "index_equity_price_T": "index_equity_price_breaks_T",
    }

    def __init__(
        self,
        reconciliation_results: Mapping[str, Mapping[str, Any]],
        date: str,
        file_path_excel: str | Path,
    ) -> None:
        self.reconciliation_results = reconciliation_results
        self.date = str(date)
        self.output_path = Path(file_path_excel)
        self.flattened_results = self._filter_and_flatten_results()
        self._export_to_excel()

    # ------------------------------------------------------------------
    # ------------------------------------------------------------------
    def _filter_and_flatten_results(self) -> Dict[Tuple[str, str, str], pd.DataFrame]:
        flat: Dict[Tuple[str, str, str], pd.DataFrame] = {}

        for date_str, fund_data in self.reconciliation_results.items():
            for fund, recon_dict in (fund_data or {}).items():
                for recon_type, subresults in (recon_dict or {}).items():
                    descriptor = RECON_DESCRIPTOR_REGISTRY.get(recon_type)
                    if not descriptor:
                        continue

                    result = self._process_from_descriptor(
                        descriptor,
                        fund,
                        date_str,
                        recon_type,
                        subresults or {},
                    )
                    if result and result.has_data():
                        flat.update(result.to_flat_dict())

                    for callback_name in descriptor.extra_callbacks:
                        callback = getattr(self, callback_name, None)
                        if callback:
                            callback(flat, fund, date_str, subresults or {}, recon_type=recon_type)
        return flat

    def _process_from_descriptor(
        self,
        descriptor: ReconDescriptor,
        fund: str,
        date_str: str,
        recon_type: str,
        subresults: Mapping[str, Any],
    ) -> ReconResult | None:
        """Build a :class:`ReconResult` from the provided descriptor."""

        holdings_df: pd.DataFrame | None = None
        if descriptor.holdings_key:
            holdings_raw = subresults.get(descriptor.holdings_key)
            if isinstance(holdings_raw, pd.DataFrame) and not holdings_raw.empty:
                holdings_df = self._build_holdings_from_descriptor(
                    descriptor,
                    holdings_raw,
                    fund,
                    date_str,
                    recon_type,
                )
        elif descriptor.require_holdings:
            # If holdings are required but no key is defined, treat as no data.
            return None

        price_t, price_t1 = self._build_price_dfs(
            descriptor,
            subresults,
            fund,
            date_str,
            recon_type,
        )

        if not any([holdings_df is not None and not holdings_df.empty, price_t is not None, price_t1 is not None]):
            if descriptor.require_holdings:
                return None

        return ReconResult(
            fund,
            date_str,
            recon_type,
            holdings_df,
            price_t,
            price_t1,
        )

    def _append_option_breakdowns(
        self,
        flat: Dict[Tuple[str, str, str], pd.DataFrame],
        fund: str,
        date_str: str,
        subresults: Mapping[str, Any],
        recon_type: str | None = None,
    ) -> None:
        regular_df = subresults.get("regular_options")
        if not regular_df.empty:
            key = (fund, date_str, "custodian_option_regular")
            flat[key] = self._prepare_holdings_df(regular_df, fund, date_str, "custodian_option_regular", "optticker")

        flex_df = subresults.get("flex_options")
        if not flex_df.empty:
            prepared = self._prepare_holdings_df(flex_df, fund, date_str, "custodian_option_flex", "optticker")
            flat[(fund, date_str, "custodian_option_flex")] = prepared
            flat[(fund, date_str, "custodian_option_flex_holdings")] = prepared

        price_t_df = subresults.get("price_discrepancies_T")
        flex_price_t, _ = split_flex_price_frames(price_t_df)
        if not flex_price_t.empty:
            prepared = self._prepare_price_df(
                flex_price_t,
                fund,
                date_str,
                "custodian_option_flex_price_T",
                "optticker",
            )
            flat[(fund, date_str, "custodian_option_flex_price_T")] = prepared

        price_t1_df = subresults.get("price_discrepancies_T1")
        flex_price_t1, _ = split_flex_price_frames(price_t1_df)
        if not flex_price_t1.empty:
            prepared = self._prepare_price_df(
                flex_price_t1,
                fund,
                date_str,
                "custodian_option_flex_price_T-1",
                "optticker",
            )
            flat[(fund, date_str, "custodian_option_flex_price_T-1")] = prepared

    def _append_option_t1_breakdowns(
        self,
        flat: Dict[Tuple[str, str, str], pd.DataFrame],
        fund: str,
        date_str: str,
        subresults: Mapping[str, Any],
    ) -> None:
        flex_df = subresults.get("flex_options")
        if not flex_df.empty:
            prepared = self._prepare_holdings_df(
                flex_df,
                fund,
                date_str,
                "custodian_option_flex_holdings_t1",
                "optticker",
            )
            flat[(fund, date_str, "custodian_option_flex_holdings_t1")] = prepared

    def _append_holdings_breakdowns(
        self,
        flat: Dict[Tuple[str, str, str], pd.DataFrame],
        fund: str,
        date_str: str,
        subresults: Mapping[str, Any],
        recon_type: str | None = None,
    ) -> None:
        final_df = subresults.get("final_recon")
        if not isinstance(final_df, pd.DataFrame) or final_df.empty:
            return

        if "breakdown" not in final_df.columns or "discrepancy_type" not in final_df.columns:
            return

        breakdowns = (
            final_df[["discrepancy_type", "breakdown"]]
            .copy()
            .assign(count=1)
            .groupby(["discrepancy_type", "breakdown"], dropna=False)["count"]
            .sum()
            .reset_index()
            .rename(columns={"count": "occurrences"})
        )

        key = f"{recon_type}_breakdowns" if recon_type else "holdings_breakdowns"
        flat[(fund, date_str, key)] = breakdowns

    def _build_holdings_from_descriptor(
        self,
        descriptor: ReconDescriptor,
        df: pd.DataFrame,
        fund: str,
        date_str: str,
        recon_type: str,
    ) -> pd.DataFrame | None:
        builder_name = descriptor.holdings_builder or "_prepare_holdings_df"
        builder: Callable[..., pd.DataFrame | None] | None = getattr(self, builder_name, None)
        if builder is None:
            return None

        ticker_col = descriptor.holdings_ticker or descriptor.price_ticker
        if not ticker_col:
            return builder(df, fund, date_str, recon_type, "TICKER")
        return builder(df, fund, date_str, recon_type, ticker_col)

    def _build_price_dfs(
        self,
        descriptor: ReconDescriptor,
        subresults: Mapping[str, Any],
        fund: str,
        date_str: str,
        recon_type: str,
    ) -> tuple[pd.DataFrame | None, pd.DataFrame | None]:
        ticker_col = descriptor.price_ticker or descriptor.holdings_ticker
        if not ticker_col:
            return None, None

        price_t: pd.DataFrame | None = None
        price_t1: pd.DataFrame | None = None
        price_t_key, price_t1_key = descriptor.price_keys

        if price_t_key:
            price_t_raw = subresults.get(price_t_key)
            price_t_raw = self._apply_price_transform(price_t_raw, descriptor)
            if not price_t_raw.empty:
                price_t = self._prepare_price_df(
                    price_t_raw,
                    fund,
                    date_str,
                    recon_type,
                    ticker_col,
                    descriptor.price_cust_col,
                )

        if price_t1_key:
            price_t1_raw = subresults.get(price_t1_key)
            price_t1_raw = self._apply_price_transform(price_t1_raw, descriptor)
            if not price_t1_raw.empty:
                price_t1 = self._prepare_price_df(
                    price_t1_raw,
                    fund,
                    date_str,
                    recon_type,
                    ticker_col,
                    descriptor.price_cust_col,
                )

        return price_t, price_t1

    def _apply_price_transform(
        self, df: pd.DataFrame | None, descriptor: ReconDescriptor
    ) -> pd.DataFrame:
        if df is None:
            return pd.DataFrame()

        if not isinstance(df, pd.DataFrame):
            return pd.DataFrame()

        if df.empty or not descriptor.price_transform:
            return df

        transform: Callable[[pd.DataFrame], pd.DataFrame] | None = getattr(
            self, descriptor.price_transform, None
        )
        if transform is None:
            return df
        transformed = transform(df)
        return transformed if isinstance(transformed, pd.DataFrame) else df

    def _select_standard_option_prices(self, df: pd.DataFrame) -> pd.DataFrame:
        _flex, standard = split_flex_price_frames(df)
        return standard

    def _prepare_index_holdings_df(
        self,
        df: pd.DataFrame,
        fund: str,
        date_str: str,
        recon_type: str,
        ticker_col: str,
    ) -> pd.DataFrame:
        df = df.copy()
        df["FUND"] = fund
        df["RECON_TYPE"] = recon_type
        df["DATE"] = date_str
        if "discrepancy_type" not in df.columns:
            df["discrepancy_type"] = "Holdings Mismatch"

        cols_to_keep = [
            col
            for col in [
                "FUND",
                "RECON_TYPE",
                "DATE",
                ticker_col,
                "discrepancy_type",
                "in_vest",
                "in_index",
            ]
            if col in df.columns
        ]
        return df[cols_to_keep] if cols_to_keep else df

    # ------------------------------------------------------------------
    def _prepare_holdings_df(
        self,
        df: pd.DataFrame,
        fund: str,
        date_str: str,
        recon_type: str,
        ticker_col: str,
    ) -> pd.DataFrame:
        df = df.copy()
        df["FUND"] = fund
        df["RECON_TYPE"] = recon_type
        df["DATE"] = date_str

        base_cols = ["FUND", "RECON_TYPE", "DATE", ticker_col, "discrepancy_type"]
        optional_cols = [
            "quantity",
            "shares_cust",
            "final_discrepancy",
            "trade_discrepancy",
            "in_vest",
            "in_cust",
            "in_index",
            "quantity_diff",
            "breakdown",
        ]
        cols_to_keep = [col for col in base_cols + optional_cols if col in df.columns]
        return df[cols_to_keep] if cols_to_keep else df

    def _limit_option_prices(self, df: pd.DataFrame, fund: str, recon_type: str) -> pd.DataFrame:
        vehicle = (FUND_DEFINITIONS.get(fund, {}).get("vehicle_wrapper") or "").lower()
        recon_lower = recon_type.lower()
        if df.empty or "option" not in recon_lower:
            return df
        if "flex" in recon_lower:
            return df
        if vehicle not in {"private_fund", "closed_end_fund"}:
            return df

        sorted_df = df.copy()
        if "option_weight" in sorted_df.columns:
            sorted_df = sorted_df.sort_values("option_weight", ascending=False)
        elif "price_diff" in sorted_df.columns:
            sorted_df = sorted_df.sort_values("price_diff", ascending=False)

        return sorted_df.head(5)

    def _prepare_price_df(
        self,
        df: pd.DataFrame,
        fund: str,
        date_str: str,
        recon_type: str,
        ticker_col: str,
        price_cust_col: str = "price_cust",
    ) -> pd.DataFrame:
        df = df.copy()
        df = self._limit_option_prices(df, fund, recon_type)
        df["FUND"] = fund
        df["RECON_TYPE"] = recon_type
        df["DATE"] = date_str
        if ticker_col in df.columns:
            df = df.rename(columns={ticker_col: "TICKER"})
        cols = [
            "FUND",
            "RECON_TYPE",
            "DATE",
            "TICKER",
            "price_vest",
            price_cust_col,
            "price_diff",
            "price_pct_diff",
            "option_weight",
            "discrepancy_type",
        ]
        cols = [c for c in cols if c in df.columns]
        return df[cols] if cols else df


    # ------------------------------------------------------------------
    def _export_to_excel(self) -> None:
        self.output_path.parent.mkdir(parents=True, exist_ok=True)
        with pd.ExcelWriter(self.output_path, engine="openpyxl") as writer:
            if not self.flattened_results:
                self._create_no_breaks_message(writer)
            else:
                self._create_summary_tab(writer)
                self._create_detailed_breaks_tab(writer)
                self._create_price_comparison_tab(writer)
                self._create_system_specific_tabs(writer)
                self._format_price_comparison_tab(writer)
            self._apply_global_header_formatting(writer)
            if writer.book.sheetnames:
                writer.book[writer.book.sheetnames[0]].sheet_state = "visible"

    def _create_no_breaks_message(self, writer: pd.ExcelWriter) -> None:
        pd.DataFrame(["All reconciliations passed - no breaks found"]).to_excel(
            writer,
            sheet_name="RECONCILIATION BREAKS SUMMARY",
            index=False,
        )

    def _create_summary_tab(self, writer: pd.ExcelWriter) -> None:
        funds = sorted(self._get_unique_funds())
        fund_data: Dict[str, Dict[str, Any]] = {}
        for fund in funds:
            row = self._build_summary_row(fund)
            row.pop("Fund", None)
            row.pop("Date", None)
            fund_data[fund] = row

        summary_df = pd.DataFrame(fund_data)
        date_row = pd.DataFrame({fund: [self.date] for fund in funds}, index=["Date"])
        summary_df = pd.concat([date_row, summary_df])
        summary_df.index.name = "Reconciliation Type"
        summary_df.to_excel(
            writer,
            sheet_name="RECONCILIATION BREAKS SUMMARY",
            index=True,
            startrow=0,
        )

        worksheet = writer.sheets["RECONCILIATION BREAKS SUMMARY"]
        has_private_or_closed = any(
            (FUND_DEFINITIONS.get(f, {}).get("vehicle_wrapper") in {"private_fund", "closed_end_fund"})
            for f in funds
        )
        if has_private_or_closed:
            note_row = len(summary_df) + 3
            cell = worksheet.cell(row=note_row, column=1)
            cell.value = (
                "Note: For private/closed-end funds, standard option price breaks show top 5 by weight. "
                "All FLEX option breaks are shown."
            )
            font = cell.font.copy(italic=True, size=9)
            cell.font = font
        worksheet.column_dimensions["A"].width = 45

    def _build_summary_row(self, fund: str) -> Dict[str, Any]:
        row = {"Fund": fund, "Date": self.date}
        columns = list(self.SUMMARY_MAPPING.values())

        for col in columns:
            row[col] = 0

        for (fund_name, _date, recon_key), df in self.flattened_results.items():
            if fund_name != fund:
                continue
            column_name = self.SUMMARY_MAPPING.get(recon_key)
            if column_name:
                row[column_name] += len(df)
        return row

    def _standardize_ticker_column(self, df: pd.DataFrame) -> pd.DataFrame:
        for col in ["eqyticker", "optticker", "occ_symbol"]:
            if col in df.columns:
                return df.rename(columns={col: "TICKER"})
        return df

    def _determine_asset_type(self, fund: Optional[str], recon_type: str, df: pd.DataFrame) -> str:
        """Infer the asset type for a reconciliation row.

        The determination prioritizes explicit recon type hints, then falls back to
        inspecting available ticker columns for option/equity/treasury cues. FLEX
        exposure is inferred either directly from the recon type or, for index funds
        flagged as FLEX, from well-known FLEX tickers in the option data.
        """
        recon_lower = recon_type.lower()
        if "flex" in recon_lower:
            return "FLEX"
        if "treasury" in recon_lower:
            return "Treasury"
        if "equity" in recon_lower:
            return "Equity"
        if "option" in recon_lower:
            return "Option"
        if "cusip" in df.columns:
            return "Treasury"
        if "eqyticker" in df.columns:
            return "Equity"
        if "optticker" in df.columns or "occ_symbol" in df.columns:
            uses_index_flex = self._uses_index_flex(fund)
            if uses_index_flex:
                return "FLEX"
            return "Option"
        return "Unknown"

    def _create_detailed_breaks_tab(self, writer: pd.ExcelWriter) -> None:
        all_breaks: list[pd.DataFrame] = []
        for (fund, date_str, recon_type), df in self.flattened_results.items():
            if "_price_" in recon_type:
                continue
            df_copy = df.copy()
            df_copy["Asset Type"] = self._determine_asset_type(fund, recon_type, df_copy)
            all_breaks.append(df_copy)
        if not all_breaks:
            return
        detailed_breaks = pd.concat(all_breaks, ignore_index=True)
        detailed_breaks = self._standardize_ticker_column(detailed_breaks)

        cols = list(detailed_breaks.columns)
        if "Asset Type" in cols:
            cols.remove("Asset Type")
            ticker_idx = next((i for i, c in enumerate(cols) if c == "TICKER"), 3)
            cols.insert(min(ticker_idx + 1, len(cols)), "Asset Type")
            detailed_breaks = detailed_breaks[cols]
        detailed_breaks.to_excel(writer, sheet_name="DETAILED BREAKS BY SECURITY", index=False)

    def _create_price_comparison_tab(self, writer: pd.ExcelWriter) -> None:
        price_comparison: list[pd.DataFrame] = []
        for (fund, date_str, recon_type), df in self.flattened_results.items():
            if "_price_" not in recon_type:
                continue
            if "t-1" in recon_type.lower():
                continue
            df_copy = df.copy()
            df_copy["FUND"] = fund
            df_copy["DATE"] = date_str
            df_copy["SOURCE"] = self._extract_source_from_recon_type(recon_type)
            df_copy["PERIOD"] = self._extract_period_from_recon_type(recon_type)
            price_comparison.append(df_copy)
        if not price_comparison:
            return
        all_prices = pd.concat(price_comparison, ignore_index=True)
        all_prices = self._limit_price_comparison_options(all_prices)
        comparison_df = self._build_price_comparison_df(all_prices)
        comparison_df.to_excel(writer, sheet_name="COMPREHENSIVE PRICE COMPARISON", index=False)

    def _limit_price_comparison_options(self, all_prices: pd.DataFrame) -> pd.DataFrame:
        if all_prices.empty or "RECON_TYPE" not in all_prices.columns:
            return all_prices

        limited_frames: list[pd.DataFrame] = []

        for (_fund, recon_type), subset in all_prices.groupby(["FUND", "RECON_TYPE"]):
            recon_lower = str(recon_type).lower()
            if "flex" in recon_lower or "option" not in recon_lower:
                limited_frames.append(subset)
                continue

            sorted_subset = subset.copy()
            if "option_weight" in sorted_subset.columns:
                sorted_subset = sorted_subset.sort_values("option_weight", ascending=False)
            elif "price_diff" in sorted_subset.columns:
                sorted_subset = sorted_subset.sort_values("price_diff", ascending=False)

            limited_frames.append(sorted_subset.head(5))

        return pd.concat(limited_frames, ignore_index=True)

    @staticmethod
    def _uses_index_flex(fund_name: Optional[str]) -> bool:
        return bool(fund_name) and fund_name in INDEX_FLEX_FUNDS

    def _build_price_comparison_df(self, all_prices: pd.DataFrame) -> pd.DataFrame:
        ticker_col = "TICKER" if "TICKER" in all_prices.columns else "equity_ticker"
        rows = []

        for (fund, ticker, period), group in all_prices.groupby(["FUND", ticker_col, "PERIOD"]):
            row = self._build_price_row(fund, ticker, period, group)
            rows.append(row)

        comparison_df = pd.DataFrame(rows)
        return comparison_df.sort_values(["Fund", "Asset Type", "Period"])

    def _build_price_row(self, fund: str, ticker: str, period: str, group: pd.DataFrame) -> Dict[str, Any]:
        """Build individual price comparison row."""
        row = {
            "Fund": fund,
            "Ticker": ticker,
            "Period": period,
            "Fund_Price": np.nan,
            "Custodian_Price": np.nan,
            "Index_Price": np.nan,
            "Price_Diff": np.nan,
            "Asset Type": self._determine_price_asset_type(group),
            "Option Weight": "",
            "% Difference": "",
        }

        # Process price data
        self._populate_price_data(row, group)

        # Calculate price difference
        prices = [p for p in [row["Fund_Price"], row["Custodian_Price"], row["Index_Price"]] if pd.notna(p)]
        if prices:
            row["Price_Diff"] = max(prices) - min(prices)

        return row

    def _determine_price_asset_type(self, group: pd.DataFrame) -> str:
        """Determine asset type from price data group."""
        if group["is_flex"].any() if "is_flex" in group.columns else False:
            return "FLEX"

        recon_types = group["RECON_TYPE"].str.lower().fillna("")
        if recon_types.str.contains("treasury").any():
            return "Treasury"
        elif recon_types.str.contains("equity").any():
            return "Equity"
        elif recon_types.str.contains("option").any():
            return "Option"
        return "Unknown"

    def _populate_price_data(self, row: Dict[str, Any], group: pd.DataFrame) -> None:
        """Populate price-related fields in the row."""
        for _, price_row in group.iterrows():
            source = str(price_row.get("SOURCE", "")).lower()

            if "custodian" in source:
                row.update({
                    "Fund_Price": price_row.get("price_vest"),
                    "Custodian_Price": price_row.get("price_cust"),
                    "Price_Diff": price_row.get("price_diff"),
                })
            elif "index" in source:
                row.update({
                    "Fund_Price": price_row.get("price_vest"),
                    "Index_Price": price_row.get("price_index", price_row.get("price_cust")),
                    "Price_Diff": price_row.get("price_diff"),
                })

            # Set optional fields
            if "option_weight" in price_row:
                weight = price_row["option_weight"]
                if isinstance(weight, (int, float)) and weight:
                    row["Option Weight"] = f"{weight * 100:.2f}%"

            if "price_pct_diff" in price_row:
                pct = price_row["price_pct_diff"]
                if isinstance(pct, (int, float)):
                    row["% Difference"] = f"{pct:.1f}%"

    def _create_system_specific_tabs(self, writer: pd.ExcelWriter) -> None:
        system_tabs = {
            "INDEX_RECONCILIATION": ["index_equity"],
            "SG_RECONCILIATION": ["sg_equity", "sg_option"],
        }
        for tab_name, recon_types in system_tabs.items():
            filtered = [df for (fund, date, recon), df in self.flattened_results.items() if recon in recon_types]
            if not filtered:
                continue
            tab_df = pd.concat(filtered, ignore_index=True)
            tab_df = self._standardize_ticker_column(tab_df)
            tab_df.to_excel(writer, sheet_name=tab_name, index=False)

    def _get_unique_funds(self) -> set[str]:
        return {fund for (fund, _date, _recon) in self.flattened_results.keys()}


    def _extract_source_from_recon_type(self, recon_type: str) -> str:
        recon_lower = recon_type.lower()
        if "custodian_equity" in recon_lower:
            return "custodian_equity"
        if "custodian_option" in recon_lower:
            return "custodian_option"
        if "custodian_treasury" in recon_lower:
            return "custodian_treasury"
        if "index_equity" in recon_lower:
            return "index_equity"
        return "unknown"

    def _extract_period_from_recon_type(self, recon_type: str) -> str:
        return "T-1" if "T-1" in recon_type else "T"

    def _format_price_comparison_tab(self, writer: pd.ExcelWriter) -> None:
        if "COMPREHENSIVE PRICE COMPARISON" not in writer.sheets:
            return

        worksheet = writer.sheets["COMPREHENSIVE PRICE COMPARISON"]
        self._apply_header_styles(worksheet)
        weight_fill = PatternFill(start_color="E2F0D9", end_color="E2F0D9", fill_type="solid")

        weight_idx = None
        pct_idx = None
        for idx, cell in enumerate(worksheet[1], start=1):
            if cell.value == "Option Weight":
                weight_idx = idx
            if cell.value == "% Difference":
                pct_idx = idx

        if weight_idx:
            col_letter = worksheet.cell(row=1, column=weight_idx).column_letter
            for row in range(2, worksheet.max_row + 1):
                cell = worksheet.cell(row=row, column=weight_idx)
                if cell.value:
                    cell.fill = weight_fill
                    cell.alignment = Alignment(horizontal="right")
        if pct_idx:
            for row in range(2, worksheet.max_row + 1):
                cell = worksheet.cell(row=row, column=pct_idx)
                if cell.value:
                    cell.alignment = Alignment(horizontal="right")

        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            worksheet.column_dimensions[column_letter].width = min(max_length + 2, 50)

    def _apply_global_header_formatting(self, writer: pd.ExcelWriter) -> None:
        for worksheet in writer.sheets.values():
            self._apply_header_styles(worksheet)

    @staticmethod
    def _apply_header_styles(worksheet) -> None:

        header_fill = PatternFill(start_color="D3E4F7", end_color="D3E4F7", fill_type="solid")
        header_font = Font(bold=True)
        for cell in worksheet[1]:
            if cell.value is None:
                continue
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")



def generate_reconciliation_reports(
    results: Mapping[str, Any],
    report_date: date | datetime | str,
    output_dir: str,
    *,
    file_name_prefix: str = "reconciliation_results",
) -> GeneratedReconciliationReport:
    normalized = normalize_reconciliation_payload(results)
    if not normalized:
        return GeneratedReconciliationReport(None)

    date_str = normalize_report_date(report_date)
    output_path = Path(output_dir)
    output_path.mkdir(parents=True, exist_ok=True)

    excel_path = output_path / f"{file_name_prefix}_{date_str}.xlsx"

    structured_details = {date_str: {}}
    for fund, payload in normalized.items():
        structured_details[date_str][fund] = payload.get("details", {})

    ReconciliationReport(structured_details, date_str, excel_path)

    return GeneratedReconciliationReport(str(excel_path))