import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import logging

logger = logging.getLogger(__name__)


def generate_trading_excel_report(comparison_data, output_path):
    """
    Generate Excel report from trading comparison data.

    Args:
        comparison_data: Dictionary with structure from TradingComplianceAnalyzer.analyze()
        output_path: Path where Excel file should be saved
    """
    import pandas as pd
    import os

    # Ensure output directory exists
    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        # 1. Summary Sheet
        summary_data = comparison_data.get('summary', {})
        summary_df = pd.DataFrame([{
            'Date': comparison_data.get('date', ''),
            'Total Funds Analyzed': summary_data.get('total_funds_analyzed', 0),
            'Funds Out of Compliance': summary_data.get('funds_out_of_compliance', 0),
            'Funds Into Compliance': summary_data.get('funds_into_compliance', 0),
            'Funds Unchanged': summary_data.get('funds_unchanged', 0),
            'Total Violations Before': summary_data.get('total_violations_before', 0),
            'Total Violations After': summary_data.get('total_violations_after', 0),
        }])
        summary_df.to_excel(writer, sheet_name='Summary', index=False)

        # 2. Fund Details Sheet
        funds_data = comparison_data.get('funds', {})

        if funds_data:
            # Create a list to hold all fund check data
            all_checks_data = []

            for fund_name, fund_info in funds_data.items():
                checks = fund_info.get('checks', {})

                for check_name, check_info in checks.items():
                    all_checks_data.append({
                        'Fund': fund_name,
                        'Compliance Check': check_name,
                        'Status Before': check_info.get('status_before', 'UNKNOWN'),
                        'Status After': check_info.get('status_after', 'UNKNOWN'),
                        'Violations Before': check_info.get('violations_before', 0),
                        'Violations After': check_info.get('violations_after', 0),
                        'Changed': 'Yes' if check_info.get('changed', False) else 'No',
                    })

            # Convert to DataFrame
            checks_df = pd.DataFrame(all_checks_data)

            # Sort by Fund, then by Compliance Check
            checks_df = checks_df.sort_values(['Fund', 'Compliance Check'])

            # Write to Excel
            checks_df.to_excel(writer, sheet_name='Compliance Details', index=False)

        # 3. Individual Fund Sheets (optional - one sheet per fund)
        for fund_name, fund_info in funds_data.items():
            checks = fund_info.get('checks', {})

            fund_check_data = []
            for check_name, check_info in checks.items():
                fund_check_data.append({
                    'Compliance Check': check_name,
                    'Status Before': check_info.get('status_before', 'UNKNOWN'),
                    'Status After': check_info.get('status_after', 'UNKNOWN'),
                    'Violations Before': check_info.get('violations_before', 0),
                    'Violations After': check_info.get('violations_after', 0),
                    'Changed': 'Yes' if check_info.get('changed', False) else 'No',
                })

            if fund_check_data:
                fund_df = pd.DataFrame(fund_check_data)
                # Excel sheet names are limited to 31 characters
                sheet_name = fund_name[:31]
                fund_df.to_excel(writer, sheet_name=sheet_name, index=False)

    print(f"Trading comparison report saved to: {output_path}")

def _create_summary_sheet(wb: Workbook, data: dict):
    """Creates executive summary sheet."""
    ws = wb.create_sheet("Executive Summary", 0)

    # Title
    ws['A1'] = "Trading Compliance Analysis - Executive Summary"
    ws['A1'].font = Font(size=14, bold=True)
    ws['A2'] = f"Date: {data['date']}"

    # Summary metrics
    summary = data['summary']
    row = 4

    metrics = [
        ("Total Funds Analyzed", summary['total_funds_analyzed']),
        ("", ""),
        ("Funds Moving OUT of Compliance", summary['funds_out_of_compliance']),
        ("Funds Moving INTO Compliance", summary['funds_into_compliance']),
        ("Funds with Unchanged Status", summary['funds_unchanged']),
        ("", ""),
        ("Total Violations (Ex-Ante)", summary['total_violations_before']),
        ("Total Violations (Ex-Post)", summary['total_violations_after']),
        ("Net Change in Violations", summary['total_violations_after'] - summary['total_violations_before'])
    ]

    for label, value in metrics:
        ws[f'A{row}'] = label
        ws[f'B{row}'] = value
        ws[f'A{row}'].font = Font(bold=True)

        # Color coding for critical metrics
        if label == "Funds Moving OUT of Compliance" and value > 0:
            ws[f'B{row}'].fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
        elif label == "Funds Moving INTO Compliance" and value > 0:
            ws[f'B{row}'].fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")

        row += 1

    # Adjust column widths
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 20


def _create_compliance_changes_sheet(wb: Workbook, data: dict):
    """Creates sheet showing compliance status changes."""
    ws = wb.create_sheet("Compliance Changes")

    # Headers
    headers = ["Fund Name", "Status Change", "Violations Before", "Violations After", "Net Change"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

    # Data rows
    row = 2
    for fund_name, fund_data in data['funds'].items():
        ws.cell(row=row, column=1, value=fund_name)
        ws.cell(row=row, column=2, value=fund_data['status_change'])
        ws.cell(row=row, column=3, value=fund_data['violations_before'])
        ws.cell(row=row, column=4, value=fund_data['violations_after'])
        ws.cell(row=row, column=5, value=fund_data['violations_after'] - fund_data['violations_before'])

        # Color code status change
        status_cell = ws.cell(row=row, column=2)
        if fund_data['status_change'] == 'OUT_OF_COMPLIANCE':
            status_cell.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
        elif fund_data['status_change'] == 'INTO_COMPLIANCE':
            status_cell.fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")

        row += 1

    # Auto-fit columns
    for col in range(1, 6):
        ws.column_dimensions[chr(64 + col)].width = 20


def _create_detailed_comparison_sheet(wb: Workbook, data: dict):
    """Creates sheet with detailed check-by-check comparison."""
    ws = wb.create_sheet("Detailed Comparison")

    # Headers
    headers = ["Fund Name", "Compliance Check", "Status Before", "Status After",
               "Violations Before", "Violations After", "Changed"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

    # Data rows
    row = 2
    for fund_name, fund_data in data['funds'].items():
        for check_name, check_data in fund_data['checks'].items():
            ws.cell(row=row, column=1, value=fund_name)
            ws.cell(row=row, column=2, value=check_name)
            ws.cell(row=row, column=3, value=check_data['status_before'])
            ws.cell(row=row, column=4, value=check_data['status_after'])
            ws.cell(row=row, column=5, value=check_data['violations_before'])
            ws.cell(row=row, column=6, value=check_data['violations_after'])
            ws.cell(row=row, column=7, value="YES" if check_data['changed'] else "NO")

            # Highlight changed rows
            if check_data['changed']:
                for col in range(1, 8):
                    ws.cell(row=row, column=col).fill = PatternFill(
                        start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"
                    )

            row += 1

    # Auto-fit columns
    for col in range(1, 8):
        ws.column_dimensions[chr(64 + col)].width = 20

