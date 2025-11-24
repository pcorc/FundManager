"""Orchestration utilities for trading compliance reporting."""

from copy import copy as copy_style
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any, Mapping, Optional

from openpyxl import load_workbook
from PyPDF2 import PdfMerger

from reporting.report_utils import normalize_report_date

from reporting.trade_compliance_analyzer import TradingComplianceAnalyzer
from reporting.trade_compliance_reporting import (
    GeneratedTradingComplianceReport,
    generate_trading_compliance_reports,
)


def build_trading_compliance_reports(
    results_ex_ante: Mapping[str, Any],
    results_ex_post: Mapping[str, Any],
    report_date: date | datetime | str,
    output_dir: str,
    *,
    traded_funds_info: Mapping[str, Mapping[str, Any]] | None = None,
    fund_registry: Mapping[str, Any] | None = None,
    file_name_prefix: str = "trading_compliance_results",
    create_pdf: bool = True,
) -> Optional[GeneratedTradingComplianceReport]:
    """Generate trading compliance comparison reports."""

    if not (results_ex_ante or results_ex_post):
        return None

    analyzer = TradingComplianceAnalyzer(
        results_ex_ante=results_ex_ante,
        results_ex_post=results_ex_post,
        date=report_date,
        traded_funds_info=traded_funds_info or {},
    )
    comparison_data = analyzer.analyze()

    return generate_trading_compliance_reports(
        comparison_data,
        report_date,
        output_dir,
        file_name_prefix=file_name_prefix,
        create_pdf=create_pdf,
    )

@dataclass
class CombinedTradingComplianceReport:
    """Aggregated Excel/PDF artefacts for trading + ex-post compliance."""

    excel_path: Optional[str]
    pdf_path: Optional[str]


def combine_trading_and_compliance_reports(
    *,
    trading_report: Optional[GeneratedTradingComplianceReport],
    compliance_report: Optional[object],
    report_date: date | datetime | str,
    output_dir: str,
    file_name_prefix: str,
) -> Optional[CombinedTradingComplianceReport]:
    """Merge trading compliance and ex-post compliance outputs into single files."""

    trading_excel = getattr(trading_report, "excel_path", None)
    compliance_excel = getattr(compliance_report, "excel_path", None)
    trading_pdf = getattr(trading_report, "pdf_path", None)
    compliance_pdf = getattr(compliance_report, "pdf_path", None)

    date_str = normalize_report_date(report_date)
    output_path = Path(output_dir)
    output_path.mkdir(parents=True, exist_ok=True)

    combined_excel = None
    if trading_excel and compliance_excel:
        combined_excel = _merge_excel_workbooks(
            Path(trading_excel),
            Path(compliance_excel),
            output_path / f"{file_name_prefix}_{date_str}.xlsx",
        )

    combined_pdf = None
    if trading_pdf and compliance_pdf:
        combined_pdf = _merge_pdfs(
            Path(trading_pdf),
            Path(compliance_pdf),
            output_path / f"{file_name_prefix}_{date_str}.pdf",
        )

    if combined_excel or combined_pdf:
        return CombinedTradingComplianceReport(combined_excel, combined_pdf)

    return None


def _merge_excel_workbooks(
    trading_path: Path, compliance_path: Path, output_path: Path
) -> Optional[str]:
    trading_wb = load_workbook(trading_path)
    compliance_wb = load_workbook(compliance_path)

    for sheet_name in compliance_wb.sheetnames:
        source = compliance_wb[sheet_name]
        _append_sheet(source, trading_wb)

    trading_wb.save(output_path)
    return str(output_path)


def _append_sheet(source_sheet, target_wb):
    target_sheet = target_wb.create_sheet(source_sheet.title)

    for row in source_sheet.iter_rows():
        for cell in row:
            target_cell = target_sheet.cell(row=cell.row, column=cell.col_idx, value=cell.value)
            if cell.has_style:
                target_cell._style = copy_style(cell._style)
            if cell.hyperlink:
                target_cell._hyperlink = copy_style(cell.hyperlink)
            if cell.comment:
                target_cell.comment = copy_style(cell.comment)

    for key, dimension in source_sheet.column_dimensions.items():
        target_dimension = target_sheet.column_dimensions[key]
        target_dimension.width = dimension.width

    for idx, dimension in source_sheet.row_dimensions.items():
        target_sheet.row_dimensions[idx].height = dimension.height

    if source_sheet.merged_cells.ranges:
        for merged_cell in source_sheet.merged_cells.ranges:
            target_sheet.merge_cells(str(merged_cell))


def _merge_pdfs(trading_pdf: Path, compliance_pdf: Path, output_path: Path) -> Optional[str]:
    merger = PdfMerger()
    try:
        for path in (trading_pdf, compliance_pdf):
            merger.append(str(path))
        merger.write(output_path)
    finally:
        merger.close()

    return str(output_path)